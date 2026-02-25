import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

new_leads = [
    # Thurgau
    {'Firma': 'Zahnteam Kreuzlingen', 'Website': 'zahnteam-kreuzlingen.ch', 'EMail': '', 'Stadt': 'Kreuzlingen', 'Telefon': '', 'WordPress': 'Ja'},
    # Dermato St. Gallen
    {'Firma': 'Dermatologie St. Gallen', 'Website': 'dermatologie-sg.ch', 'EMail': '', 'Stadt': 'St. Gallen', 'Telefon': '', 'WordPress': 'Ja'},
    # Dermato Winterthur
    {'Firma': 'Hautstation Winterthur', 'Website': 'hautstation.ch', 'EMail': '', 'Stadt': 'Winterthur', 'Telefon': '', 'WordPress': 'Ja'},
    # Ortho Winterthur
    {'Firma': 'OZW Orthopädisches Zentrum', 'Website': 'ozw.ch', 'EMail': '', 'Stadt': 'Winterthur', 'Telefon': '+41 52 203 40 05', 'WordPress': 'Ja'},
    # Ortho Zürich
    {'Firma': 'Sportorthopäde Zürich', 'Website': 'sportorthopaede.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '', 'WordPress': 'Ja'},
    # Zahnarzt Aarau
    {'Firma': 'Zahnästhetik Aarau', 'Website': 'aarau-zahnaesthetik.ch', 'EMail': '', 'Stadt': 'Aarau', 'Telefon': '+41 62 823 04 33', 'WordPress': 'Ja'},
    # Augenarzt Zug
    {'Firma': 'HEAL Augenpraxis Zug', 'Website': 'augenarztzug.ch', 'EMail': 'info@augenarztzug.ch', 'Stadt': 'Zug', 'Telefon': '+41 41 711 70 56', 'WordPress': 'Ja'},
    # Augenarzt Bern
    {'Firma': 'Augenärzte Zentrum Marktgasse', 'Website': 'augenaerzte-bern.ch', 'EMail': '', 'Stadt': 'Bern', 'Telefon': '+41 31 310 10 00', 'WordPress': 'Ja'},
    # Hausarzt Solothurn
    {'Firma': 'Gemeinschaftspraxis Dornacherplatz', 'Website': 'gemeinschaftspraxis-dornacherplatz.ch', 'EMail': '', 'Stadt': 'Solothurn', 'Telefon': '+41 32 672 10 23', 'WordPress': 'Ja'},
]

def normalize_url(url):
    url = str(url).strip().lower().rstrip('/')
    url = url.replace('https://', '').replace('http://', '').replace('www.', '')
    return url

wb = openpyxl.load_workbook(input_file)
ws = wb['Praxen Leads']
existing_urls = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    val = str(row[0].value or '').strip()
    if val: existing_urls.add(normalize_url(val))
print(f"Bestehende Eintraege: {ws.max_row - 1}")

thin_border = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'),
                     top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))
data_font = Font(name='Calibri', size=11, color='333333')
data_alignment = Alignment(horizontal='left', vertical='center')
added = skipped = 0
for lead in new_leads:
    norm_url = normalize_url(lead['Website'])
    if norm_url in existing_urls:
        print(f"  SKIP: {lead['Website']}"); skipped += 1; continue
    next_row = ws.max_row + 1
    for i, key in enumerate(['Firma','Website','EMail','Stadt','Telefon','WordPress'], 1):
        cell = ws.cell(row=next_row, column=i, value=lead[key])
        cell.font = data_font; cell.alignment = data_alignment; cell.border = thin_border
    existing_urls.add(norm_url); added += 1
    print(f"  ADDED: {lead['Firma']} ({lead['Website']})")
wb.save(input_file)
print(f"\nErgebnis: {added} hinzugefuegt, {skipped} uebersprungen")
print(f"Gesamt 'Praxen Leads': {ws.max_row - 1} Eintraege")
