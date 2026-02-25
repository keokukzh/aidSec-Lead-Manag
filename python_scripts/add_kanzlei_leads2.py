import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

new_leads = [
    {'Firma': 'BUCHLI JUST', 'Website': 'churanwalt.ch', 'EMail': '', 'Stadt': 'Chur', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Anwaltskanzlei Oesch', 'Website': 'anwaltskanzlei-chur.ch', 'EMail': 'laura.oesch@oesch-law.ch', 'Stadt': 'Chur', 'Telefon': '+41 81 515 57 57', 'WordPress': 'Ja'},
    {'Firma': 'Frôté & Partner', 'Website': 'frotepartner.ch', 'EMail': 'info@frotepartner.ch', 'Stadt': 'Biel', 'Telefon': '+41 32 322 25 21', 'WordPress': 'Ja'},
    {'Firma': 'Kaiser Odermatt & Partner', 'Website': 'kaiserodermatt.ch', 'EMail': '', 'Stadt': 'Zug', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Schmidhäusler Rechtsanwälte AG', 'Website': 'schmidhaeusler.ch', 'EMail': 'anwalt@schmidhaeusler.ch', 'Stadt': 'Galgenen', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Streiff Rechtsanwälte AG', 'Website': 'this-law.ch', 'EMail': 'info@this-law.ch', 'Stadt': 'Wetzikon', 'Telefon': '+41 44 932 15 09', 'WordPress': 'Ja'},
    {'Firma': 'Lehner Trüeb Küng', 'Website': 'advokatenliestal.ch', 'EMail': '', 'Stadt': 'Liestal', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Advokatur Polatli', 'Website': 'advokatur-polatli.ch', 'EMail': '', 'Stadt': 'Liestal', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Rüesch Rechtsanwälte', 'Website': 'sglaw.ch', 'EMail': '', 'Stadt': 'St. Gallen', 'Telefon': '+41 71 227 30 30', 'WordPress': 'Ja'},
    {'Firma': 'LENLAW', 'Website': 'len.law', 'EMail': 'info@len.law', 'Stadt': 'Bern', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Anwaltskanzlei Eshrefi', 'Website': 'anwaltskanzlei-eshrefi.ch', 'EMail': 'info@anwaltskanzlei-eshrefi.ch', 'Stadt': 'Winterthur', 'Telefon': '+41 78 214 15 45', 'WordPress': 'Ja'},
    {'Firma': 'Grunder Rechtsanwälte AG', 'Website': 'grunder-law.ch', 'EMail': 'info@grunder-law.ch', 'Stadt': 'Baar', 'Telefon': '+41 41 500 31 00', 'WordPress': 'Ja'},
    {'Firma': 'FD Anwaltskanzlei AG', 'Website': 'fd-anwaltskanzlei.ch', 'EMail': 'flavia.dudler@fd-anwaltskanzlei.ch', 'Stadt': 'Bülach', 'Telefon': '+41 44 860 38 15', 'WordPress': 'Ja'},
    {'Firma': 'Dietsche Rechtsanwälte & Notare', 'Website': 'dietsche-law.ch', 'EMail': '', 'Stadt': 'Rorschach', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Höfliger Rechtsanwälte AG', 'Website': 'eh-law.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Schlatter Aepli Partner', 'Website': 'schlatter-aepli.ch', 'EMail': 'office@schlatter-aepli.ch', 'Stadt': 'Kreuzlingen', 'Telefon': '+41 71 677 97 87', 'WordPress': 'Ja'},
    {'Firma': 'Gruber & Gattlen', 'Website': 'anwalt-wallis.ch', 'EMail': 'david.gruber@anwalt-wallis.ch', 'Stadt': 'Visp', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Meier Vogel Partner GmbH', 'Website': 'mvlaw.ch', 'EMail': 'info@mvlaw.ch', 'Stadt': 'Dübendorf', 'Telefon': '+41 44 820 20 17', 'WordPress': 'Ja'},
]


def normalize_url(url):
    url = str(url).strip().lower().rstrip('/')
    url = url.replace('https://', '').replace('http://', '').replace('www.', '')
    return url


wb = openpyxl.load_workbook(input_file)
ws = wb["Anwalts Kanzleien"]

existing_urls = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    url = normalize_url(row[0].value or '')
    if url:
        existing_urls.add(url)

print(f"Bestehende Eintraege: {len(existing_urls)}")

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
data_font = Font(name='Calibri', size=11)
data_alignment = Alignment(vertical='center', wrap_text=False)

added = 0
skipped = 0

for lead in new_leads:
    url_norm = normalize_url(lead['Website'])
    if url_norm in existing_urls:
        print(f"  SKIP (Duplikat): {lead['Firma']} ({lead['Website']})")
        skipped += 1
        continue

    next_row = ws.max_row + 1
    col_map = {1: 'Firma', 2: 'Website', 3: 'EMail', 4: 'Stadt', 5: 'Telefon', 6: 'WordPress'}

    for col_idx, field in col_map.items():
        cell = ws.cell(row=next_row, column=col_idx, value=lead.get(field, ''))
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

    existing_urls.add(url_norm)
    added += 1
    print(f"  ADD: {lead['Firma']} ({lead['Website']}) - {lead['Stadt']}")

print(f"\nErgebnis: {added} neue Leads hinzugefuegt, {skipped} Duplikate uebersprungen")
print(f"Gesamtanzahl: {ws.max_row - 1} Eintraege")

wb.save(input_file)
print(f"Gespeichert: {input_file}")
