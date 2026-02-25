import openpyxl

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"
wb = openpyxl.load_workbook(input_file)
ws = wb["Anwalts Kanzleien"]

fixes = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
    for cell in row:
        if cell.value is not None and str(cell.value).strip().lower() == 'nan':
            cell.value = ''
            fixes += 1
        if cell.value is not None and str(cell.value).strip() == '':
            cell.value = None
            fixes += 1

# Row-level fixes based on website column (col B = index 1)
for row_idx in range(2, ws.max_row + 1):
    website = str(ws.cell(row=row_idx, column=2).value or '').lower()

    # RÄWEL ADVOKATUR was removed as duplicate, so row 1 should be the entry with www.
    if 'raewel-advokatur' in website:
        ws.cell(row=row_idx, column=1).value = 'RÄWEL ADVOKATUR'

    if 'badertscher' in str(ws.cell(row=row_idx, column=1).value or '').lower() or 'b-legal' in website:
        ws.cell(row=row_idx, column=1).value = 'Badertscher Rechtsanwälte AG'

print(f"Fixes angewendet: {fixes}")

# Verify final state
print(f"\nFinale Validierung:")
missing = {'Firma': 0, 'EMail': 0, 'Stadt': 0, 'Telefon': 0}
total = 0
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
    total += 1
    firma = row[0].value
    email = row[2].value
    stadt = row[3].value
    tel = row[4].value
    if not firma or str(firma).strip() == '':
        missing['Firma'] += 1
        print(f"  Firma fehlt: Zeile {row[0].row}")
    if not email or str(email).strip() == '':
        missing['EMail'] += 1
    if not stadt or str(stadt).strip() == '':
        missing['Stadt'] += 1
        print(f"  Stadt fehlt: Zeile {row[0].row} - {firma} ({row[1].value})")
    if not tel or str(tel).strip() == '':
        missing['Telefon'] += 1

print(f"\nGesamt: {total} Einträge")
print(f"Firmennamen:  {total - missing['Firma']}/{total}")
print(f"E-Mail:       {total - missing['EMail']}/{total}")
print(f"Stadt:        {total - missing['Stadt']}/{total}")
print(f"Telefon:      {total - missing['Telefon']}/{total}")

wb.save(input_file)
print(f"\nDatei gespeichert.")
