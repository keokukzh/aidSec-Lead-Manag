import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

new_leads = [
    # Familienrecht / Scheidung
    {'Firma': 'divortis AG', 'Website': 'divortis.ch', 'EMail': '', 'Stadt': 'Basel', 'Telefon': '+41 61 206 45 75', 'WordPress': 'Ja'},
    {'Firma': 'Advokaturbüro Scheidungen', 'Website': 'rechtsanwalt-scheidungen.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Baudenbacher Law AG', 'Website': 'baudenbacher.law', 'EMail': 'info@baudenbacher.law', 'Stadt': 'Zürich', 'Telefon': '+41 44 562 70 27', 'WordPress': 'Ja'},

    # Arbeitsrecht
    {'Firma': 'Rechtsanwältin Regula Bärtschi', 'Website': 'anwaeltin-zuerich.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'entlassung.ch', 'Website': 'entlassung.ch', 'EMail': '', 'Stadt': '', 'Telefon': '', 'WordPress': 'Ja'},

    # Gesellschaftsrecht
    {'Firma': 'ComLaw', 'Website': 'comlaw.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '', 'WordPress': 'Ja'},

    # Datenschutz
    {'Firma': 'Kanzlei Berger GmbH', 'Website': 'kanzlei-berger.ch', 'EMail': 'denis@kanzlei-berger.ch', 'Stadt': 'Biel', 'Telefon': '+41 79 473 61 25', 'WordPress': 'Ja'},

    # Steuerrecht
    {'Firma': 'Tax Partner AG', 'Website': 'taxpartner.ch', 'EMail': 'reception@taxpartner.ch', 'Stadt': 'Zürich', 'Telefon': '+41 44 215 77 77', 'WordPress': 'Ja'},
    {'Firma': 'LMV Partner AG', 'Website': 'lmvpartner.ch', 'EMail': '', 'Stadt': 'Zug', 'Telefon': '', 'WordPress': 'Ja'},

    # Inkasso / Vertragsrecht
    {'Firma': 'Dr. Fassbender Rechtsanwälte', 'Website': 'dr-fassbender.com', 'EMail': '', 'Stadt': '', 'Telefon': '', 'WordPress': 'Ja'},

    # Verkehrsrecht
    {'Firma': 'Basel Advokaten', 'Website': 'basel-advokaten.ch', 'EMail': '', 'Stadt': 'Basel', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Rechtsanwalt Geier', 'Website': 'rechtsanwalt-geier.ch', 'EMail': '', 'Stadt': 'Thalwil', 'Telefon': '', 'WordPress': 'Ja'},

    # Regionale Lücken
    {'Firma': 'Anwaltskanzlei Marcel Furrer', 'Website': 'anwaltcham.ch', 'EMail': '', 'Stadt': 'Cham', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'LINDTLAW', 'Website': 'lindtlaw.ch', 'EMail': 'office@lindtlaw.ch', 'Stadt': 'Kreuzlingen', 'Telefon': '+41 71 678 10 10', 'WordPress': 'Ja'},
]

def normalize_url(url):
    url = str(url).strip().lower().rstrip('/')
    url = url.replace('https://', '').replace('http://', '').replace('www.', '')
    return url

wb = openpyxl.load_workbook(input_file)
ws = wb['Anwalts Kanzleien']

existing_urls = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    val = str(row[0].value or '').strip()
    if val:
        existing_urls.add(normalize_url(val))

print(f"Bestehende Eintraege: {len(existing_urls)}")

thin_border = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0')
)
data_font = Font(name='Calibri', size=11, color='333333')
data_alignment = Alignment(horizontal='left', vertical='center')

added = 0
skipped = 0

for lead in new_leads:
    norm_url = normalize_url(lead['Website'])
    if norm_url in existing_urls:
        print(f"  SKIP: {lead['Website']}")
        skipped += 1
        continue

    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=lead['Firma'])
    ws.cell(row=next_row, column=2, value=lead['Website'])
    ws.cell(row=next_row, column=3, value=lead['EMail'])
    ws.cell(row=next_row, column=4, value=lead['Stadt'])
    ws.cell(row=next_row, column=5, value=lead['Telefon'])
    ws.cell(row=next_row, column=6, value=lead['WordPress'])

    for col in range(1, 7):
        cell = ws.cell(row=next_row, column=col)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

    existing_urls.add(norm_url)
    added += 1
    print(f"  ADDED: {lead['Firma']} ({lead['Website']})")

wb.save(input_file)
print(f"\nErgebnis: {added} hinzugefuegt, {skipped} uebersprungen")
print(f"Gesamt 'Anwalts Kanzleien': {ws.max_row - 1} Eintraege")
