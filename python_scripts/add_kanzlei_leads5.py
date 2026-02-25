import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

new_leads = [
    # V1: Google Dorks
    {'Firma': '4Sight Legal', 'Website': '4sightlegal.ch', 'EMail': 'manuel.bertschi@4sightlegal.ch', 'Stadt': 'Zürich', 'Telefon': '+41 44 442 40 57', 'WordPress': 'Ja'},
    {'Firma': 'GMP Legal', 'Website': 'gmplegal.ch', 'EMail': 'info@gmplegal.com', 'Stadt': 'Zürich', 'Telefon': '+41 43 255 10 00', 'WordPress': 'Ja'},
    {'Firma': 'rtwp rechtsanwälte & notare', 'Website': 'rtwp.ch', 'EMail': 'info@rtwp.ch', 'Stadt': 'St. Gallen', 'Telefon': '+41 71 228 70 00', 'WordPress': 'Ja'},
    {'Firma': 'GÖRG SUTER AG', 'Website': 'goergsuter.ch', 'EMail': '', 'Stadt': 'Staad SG', 'Telefon': '+41 41 712 30 50', 'WordPress': 'Ja'},
    {'Firma': 'Martin Rechtsanwälte GmbH', 'Website': 'martin-ra.ch', 'EMail': '', 'Stadt': 'Rorschach', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'WS LAW Winzeler Steffen', 'Website': 'winzelersteffen.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Schwegler & Partner', 'Website': 'schwep.ch', 'EMail': '', 'Stadt': 'Laupen', 'Telefon': '', 'WordPress': 'Ja'},

    # V2: Regionale Lücken
    {'Firma': 'Beaudouin Advokatur', 'Website': 'beaudouin-advokatur.ch', 'EMail': 'info@beaudouin-advokatur.ch', 'Stadt': 'Stans', 'Telefon': '+41 41 610 11 44', 'WordPress': 'Ja'},
    {'Firma': 'Blöchlinger Iten Fessler', 'Website': 'bilaw.ch', 'EMail': '', 'Stadt': 'Stans', 'Telefon': '+41 41 618 62 00', 'WordPress': 'Ja'},
    {'Firma': 'Bäni Advokatur und Mediation', 'Website': 'baeni.org', 'EMail': 'mail@baeni.org', 'Stadt': 'Seewen SZ', 'Telefon': '+41 41 543 62 03', 'WordPress': 'Ja'},
    {'Firma': 'Bruhin Klass Landtwing', 'Website': 'bkl-recht.ch', 'EMail': '', 'Stadt': 'Schwyz', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Lüönd Legal', 'Website': 'luondlegal.ch', 'EMail': '', 'Stadt': 'Einsiedeln', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Lanz Wehrli Advokatur AG', 'Website': 'lanzwehrli.ch', 'EMail': '', 'Stadt': 'Zofingen', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'SLP Rechtsanwälte und Notariat', 'Website': 'slp.ch', 'EMail': '', 'Stadt': 'Aarau', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'BALEX AG', 'Website': 'balex.law', 'EMail': '', 'Stadt': 'Binningen', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Anwalt Pedolin', 'Website': 'schweizer-rechtsanwalt.com', 'EMail': '', 'Stadt': 'Langrickenbach', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Wehrli Partner Rechtsanwälte', 'Website': 'wehrlipartner.ch', 'EMail': '', 'Stadt': 'Aargau', 'Telefon': '', 'WordPress': 'Ja'},

    # V5: Nischen
    {'Firma': 'E. Blum & Co. AG', 'Website': 'eblum.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Troesch Scheidegger Werner AG', 'Website': 'tswpat.ch', 'EMail': '', 'Stadt': 'Zumikon', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Staedeli Legal Partners', 'Website': 'staedeli-legalpartners.ch', 'EMail': '', 'Stadt': '', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Kleb & Partner Rechtsanwälte', 'Website': 'klebundpartner.ch', 'EMail': '', 'Stadt': 'Wollerau', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Ramseier Anwaltskanzlei', 'Website': 'ramseierrecht.ch', 'EMail': '', 'Stadt': '', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Schütz Law AG', 'Website': 'schuetzlaw.ch', 'EMail': '', 'Stadt': '', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'ViciMed AG', 'Website': 'vicimed.ch', 'EMail': '', 'Stadt': '', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'CORE Attorneys', 'Website': 'core-attorneys.com', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Roman Baumann Lorant', 'Website': 'stiftungen-vereine.ch', 'EMail': '', 'Stadt': 'Basel', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'DUFOUR Advokatur AG', 'Website': 'dufour-advokatur.ch', 'EMail': '', 'Stadt': 'Basel', 'Telefon': '', 'WordPress': 'Ja'},
]

def normalize_url(url):
    url = str(url).strip().lower().rstrip('/')
    url = url.replace('https://', '').replace('http://', '').replace('www.', '')
    return url

wb = openpyxl.load_workbook(input_file)
ws = wb['Anwalts Kanzleien']

existing_urls = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    val = str(row[0].value or '').strip()
    if val:
        existing_urls.add(normalize_url(val))

print(f"Bestehende Eintraege: {len(existing_urls)}")

thin_border = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0')
)
data_font = Font(name='Calibri', size=11, color='333333')
data_alignment = Alignment(horizontal='left', vertical='center')

added = 0
skipped = 0

for lead in new_leads:
    norm_url = normalize_url(lead['Website'])
    if norm_url in existing_urls:
        print(f"  SKIP: {lead['Website']}")
        skipped += 1
        continue

    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=lead['Firma'])
    ws.cell(row=next_row, column=2, value=lead['Website'])
    ws.cell(row=next_row, column=3, value=lead['EMail'])
    ws.cell(row=next_row, column=4, value=lead['Stadt'])
    ws.cell(row=next_row, column=5, value=lead['Telefon'])
    ws.cell(row=next_row, column=6, value=lead['WordPress'])

    for col in range(1, 7):
        cell = ws.cell(row=next_row, column=col)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

    existing_urls.add(norm_url)
    added += 1
    print(f"  ADDED: {lead['Firma']} ({lead['Website']})")

wb.save(input_file)
print(f"\nErgebnis: {added} hinzugefuegt, {skipped} uebersprungen")
print(f"Gesamt 'Anwalts Kanzleien': {ws.max_row - 1} Eintraege")
