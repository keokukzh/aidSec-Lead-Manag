import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

# Read existing data
existing_df = pd.read_excel(input_file)
existing_websites = existing_df['Website'].str.strip().str.lower().tolist()

# New leads data - parsed from user message
new_leads_raw = """Hausarzt-Praxis Dr. Dreiding,dreiding.ch,,Zürich
Hausarzt Enge,hausarzt-enge.ch,hausarzt-enge@hin.ch,Zürich
Arzthaus Zürich City,arzthaus.ch,zurich@arzthaus.ch,Zürich
Praxis Albisrieder Dörfli,praxis-albisriederdoerfli.ch,praxis-albisriederdoerfli@hin.ch,Zürich
Praxis am Steinwiesplatz,praxisamsteinwiesplatz.ch,praxisamsteinwiesplatz@hin.ch,Zürich
Praxis Dr. Ing,praxis-dr-ing.ch,praxis.ing@hin.ch,Zürich
Praxis Almut Meyer,praxis-almut-meyer.ch,praxis.meyer@hin.ch,Zürich
Praxis Dr. Zahiri,praxisdrzahiri.ch,info@praxisdrzahiri.ch,Zürich
Hausärztin Akupunktur,hausaerztin-akupunktur.ch,ursula.scharf@hin.ch,Zürich
Truong Medical,truong-medical.ch,kontakt@truong-medical.ch,Zürich
Arztpraxis Hittnau,arztpraxis-hittnau.ch,info@arztpraxis-hittnau.ch,Zürich
Hausärzte Witikon,hausaerzte-witikon.ch,hausaerzte-zentrum-witikon@hin.ch,Zürich
Arztpraxis Kalkbreite,arztpraxiskalkbreite.ch,kontakt@arztpraxiskalkbreite.ch,Zürich
Praxis Zähringer Ahornweg,praxis-zaehringer-ahornweg.ch,praxis.zaehringer@hin.ch,Bern
Zeughauspraxis,zeughauspraxis.ch,zeughauspraxis@hin.ch,Bern
PraxisPlus,praxisplus.ch,praxisplus@hin.ch,Bern
BaselMed,baselmed.ch,info@baselmed.ch,Basel
Hausärzte Basel,hausaerzte-basel.ch,hausaerzte-basel@hin.ch,Basel
SoC in22,socin22.ch,,Basel
Rüti Meier,ruetimeyer.ch,praxis.dubach@hin.ch,Basel
MedCenter Volta,medcentervolta.ch,medcentervolta@hin.ch,Basel
Hausarztpraxis Barac,hausarztpraxis-barac.ch,info@praxisbarac.ch,Basel
Pilatus Praxis,pilatuspraxis.ch,info@pilatuspraxis.ch,Luzern
Hausarztpraxis Würzenbach,hausarztpraxis-wuerzenbach.ch,info@hausarztpraxis-wuerzenbach.ch,Luzern
Hausarzt Dr. Wüst,hausarzt-drwuest.ch,info@praxisdrwuest.ch,Luzern
Mzemmen,mzemmen.ch,praxis@mzemmen.ch,Luzern
Praxis Lerchental,praxis-lerchental.ch,praxis-lerchental@friendlydocs.ch,St. Gallen
Medizin am Park,medizin-am-park.ch,info@medizin-am-park.ch,St. Gallen
Hausarzt Poststrasse,hausarztpoststrasse.ch,mrcs@hin.ch,St. Gallen
Praxis Schuster,praxis-schuster.ch,,St. Gallen
Eglizahnmedizin,eglizahnmedizin.ch,info@eglizahnmedizin.ch,St. Gallen
Zahnarztpraxis Buchegg,zahnarztpraxis-buchegg.ch,info@zahnarztpraxis-buchegg.ch,Zürich
Albisrieden Zahnarzt,albisrieden-zahnarzt.ch,praxis@albisrieden-zahnarzt.ch,Zürich
Zahnarzt Bux,zahnarztbux.ch,praxis@zahnarztbux.ch,Zürich
Zahn Gemeinschaftspraxis,zahn-gemeinschaftspraxis.ch,,Zürich
Zahnerzte Am Bleicherweg,zahnaerzteambleicherweg.ch,info@zahnaerzteambleicherweg.ch,Zürich
Zahnarzt Schumacher,zahnarztschumacher.ch,empfang@zahnarztschumacher.ch,Zürich
Zahnarzt Kyburz,zahnarzt-kyburz.ch,praxis@zahnarzt-kyburz.ch,Zürich
Parkring 8,parkring8.ch,dentist@parkring8.ch,Zürich
Zahnoz,zahnoz.ch,info@zahnoz.ch,Zürich
Schütz Dental,schuetzdental.ch,,Zürich
Zürich Dental,zurichdental.ch,info@zurichdental.ch,Zürich
Zahnarzt Praxis ZZG,zahnarztpraxis-zuerich-zzg.ch,info@zzg.ch,Zürich
Dentalarzt,dentalarzt.ch,info@dentalarzt.ch,Zürich
Zahn Loft,zahnloft.ch,praxis@zahnloft.ch,Zürich
Zahnarzt Schneebeli,zahnarztschneebeli.ch,,Zürich
Zahir Physio,zahir-physio.ch,,Zürich
Physio Hofmann,physio-hofmann.ch,info@physio-hofmann.ch,Zürich
Physio Dynamics,physio-dynamics.ch,,Zürich
Physio Zentrum,physiozentrum.ch,,Zürich
Theracare 360,theracare360.ch,info@theracare360.ch,Zürich
Physio SGC,physio-sgc.ch,,Zürich
Rösch Eisen,roescheisen.ch,praxis@roescheisen.ch,Zürich
Limmatz,limmatz.ch,info@limmatz.ch,Zürich
Augenpraxis Odeon,augenpraxisodeon.ch,info@augenpraxisodeon.ch,Zürich
Augen Arzt Praxis,augen-arzt-praxis.ch,hallo@augen-arzt-praxis.ch,Zürich
Sanaskin,sanaskin.ch,,Zürich
Skinmed Zürich,skinmed.ch,,Zürich
Derma Medical Clinic,dermamedicalclinic.ch,info@dermamedicalclinic.ch,Zürich
Unsere Haut,unsere-haut.ch,,Zürich
Derma Autoquai,dermautoquai.ch,info@medizeplin.de,Zürich
Orto Fuss,orthofuss.ch,info@orthofuss.ch,Zürich
OSZH,oszh.ch,praxis@oszh.ch,Zürich
Zotz,zotz.ch,meilen@zotz.ch,Zürich
Orthopädie Wüst,orthopaedie-wuest.ch,info@orthopaedie-wuest.ch,Zürich"""

from io import StringIO
new_leads_df = pd.read_csv(StringIO(new_leads_raw), header=None, names=['Firma', 'Website', 'EMail', 'Stadt'])
new_leads_df['Telefon'] = ''

# Normalize websites - add https:// if missing
def normalize_website(url):
    url = str(url).strip()
    if url and not url.startswith('http'):
        url = 'https://' + url
    return url

new_leads_df['Website'] = new_leads_df['Website'].apply(normalize_website)

# Check for duplicates
new_leads_df['Website_clean'] = new_leads_df['Website'].str.strip().str.lower()
duplicates_mask = new_leads_df['Website_clean'].isin(existing_websites)

duplicates_count = duplicates_mask.sum()
print(f"Vorhandene Einträge: {len(existing_df)}")
print(f"Duplikate gefunden: {duplicates_count}")

# Show duplicates
if duplicates_count > 0:
    print("\nDuplikate (werden übersprungen):")
    for idx, row in new_leads_df[duplicates_mask].iterrows():
        print(f"  - {row['Firma']}")

# Filter only new leads
new_only = new_leads_df[~duplicates_mask].copy()
new_only = new_only[['Firma', 'Website', 'EMail', 'Stadt', 'Telefon']]

print(f"Neue Leads hinzugefügt: {len(new_only)}")

# Combine
combined = pd.concat([existing_df, new_only], ignore_index=True)

# Save with formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Praxen Leads"

for r_idx, row in enumerate(dataframe_to_rows(combined, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20

thin_border = Border(bottom=Side(style='thin', color='000000'))
for cell in ws[1]:
    cell.border = thin_border

wb.save(input_file)
print(f"\nDatei aktualisiert: {input_file}")
print(f"Gesamtanzahl Einträge: {len(combined)}")
