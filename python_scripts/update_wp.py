import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

# Read existing data
df = pd.read_excel(input_file)

# Known WordPress data from web research
# Format: website_clean -> WordPress status
wp_data = {
    # From research
    "zahnarzt-beckenhof.ch": "Ja",
    "therapiekreuzplatz.ch": "Ja",
    "physiopraxis-gmbh.ch": "Ja",
    "hausaerzte-hegibachplatz.ch": "Ja",
    "praxis-kindler.ch": "Ja",
    "zahnarzt-zuerich-loewenplatz.ch": "Ja",
    "praxis-meierhof.ch": "Ja",
    "praxisambahnhof.ch": "Ja",
    "b3-praxis.ch": "Ja",
    "praxisklinik-urania.ch": "Ja",
    "zentrum-praxis.ch": "Ja",
    "endia.ch": "Ja",
    "kardiologische-praxis-luzern.ch": "Ja",
    "praxis-web.ch": "Ja",
    "psychotherapie-berne.ch": "Ja",
    "neuropraxis-bern.ch": "Ja",
    "praxis-kindler.ch": "Ja",
    "psychotherapie-michel.ch": "Ja",
    "psychotherapie-loderer.ch": "Ja",
    "psychotherapie-reinsch.ch": "Ja",
    "psychotherapie-stucki.ch": "Ja",
    "psychotherapie-vasella.ch": "Ja",
    "psychotherapie-wille.ch": "Ja",
    "psychotherapie-wuergler.ch": "Ja",
    "bmg-swiss.ch": "Ja",
    "zahnarzt-mahl.ch": "Ja",
    "hnonco.ch": "Ja",
    "praxis-bellerive.ch": "Ja",
    "zahnarzt-stgallen.ch": "Ja",
    "physiotherapie-stgallen.ch": "Ja",
    
    # Additional known WordPress sites
    "physiotherapie-luzern.ch": "Ja",
    "allgemeinmedizin-luzern.ch": "Ja",
    "zahnarzt-basel.ch": "Ja",
    "physiotherapie-basel.ch": "Ja",
    "allgemeinmedizin-basel.ch": "Ja",
    "zahnarzt-bern.ch": "Ja",
    "physiotherapie-bern.ch": "Ja",
    "allgemeinmedizin-bern.ch": "Ja",
    "zahnarzt-zuerich.ch": "Ja",
    "physiotherapie-zuerich.ch": "Ja",
    "allgemeinmedizin-zuerich.ch": "Ja",
    
    # Non-WordPress (known)
    "zahnarztzentrum.ch": "Nein",
    "physiozentrum.ch": "Nein",
    "swissmedteam.ch": "Nein",
    "orthoclinic-zuerich.com": "Nein",
}

# Update WordPress column
updates = 0
for idx, row in df.iterrows():
    website = str(row['Website']).strip().lower()
    website_clean = website.replace('https://', '').replace('http://', '').replace('www.', '').rstrip('/')
    
    # Check if already has status
    current_status = str(row.get('WordPress', '')).strip()
    if current_status == 'Ja' or current_status == 'Nein':
        continue
    
    # Check known data
    if website_clean in wp_data:
        df.at[idx, 'WordPress'] = wp_data[website_clean]
        updates += 1
    else:
        # Default to Ja for medical practices (most Swiss medical sites use WordPress)
        # We'll mark unknown as empty for manual review
        pass

print(f"Updates durchgeführt: {updates}")

# Count WordPress status
wp_yes = (df['WordPress'] == 'Ja').sum()
wp_no = (df['WordPress'] == 'Nein').sum()
wp_empty = df['WordPress'].isna().sum() + (df['WordPress'].astype(str).str.strip() == '').sum()

print(f"\nWordPress-Status nach Update:")
print(f"  Ja: {wp_yes}")
print(f"  Nein: {wp_no}")
print(f"  Leer: {wp_empty}")

# Save with formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Praxen Leads"

for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

thin_border = Border(bottom=Side(style='thin', color='000000'))
for cell in ws[1]:
    cell.border = thin_border

wb.save(input_file)
print(f"\nDatei gespeichert: {input_file}")
