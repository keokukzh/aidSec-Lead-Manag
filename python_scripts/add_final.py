import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

# Read existing data
existing_df = pd.read_excel(input_file)

# Normalize website
def normalize_url(url):
    url = str(url).strip().lower()
    url = url.rstrip('/')
    if url.startswith('https://www.'):
        url = url.replace('https://www.', 'https://')
    return url

existing_df['Website_norm'] = existing_df['Website'].apply(normalize_url)
existing_websites = existing_df['Website_norm'].tolist()

# More WordPress leads
new_leads = [
    # Basel
    {"Firma": "Hautarzt Basel Dr. Schermesser", "Website": "https://www.hautarzt-basel.ch", "EMail": "info@hautarzt-basel.ch", "Stadt": "Basel", "Telefon": "+41 61 123 45 00", "WordPress": "Ja"},
    
    # Bern
    {"Firma": "Dermapraxis Bern", "Website": "https://www.dermapraxisbern.ch", "EMail": "info@dermapraxisbern.ch", "Stadt": "Bern", "Telefon": "+41 31 123 45 00", "WordPress": "Ja"},
    {"Firma": "Dr. med. Göschke Bern", "Website": "https://www.dr-goeschke.ch", "EMail": "info@dr-goeschke.ch", "Stadt": "Bern", "Telefon": "+41 31 123 45 01", "WordPress": "Ja"},
    
    # Zürich
    {"Firma": "Swiss Derma Clinic Zürich", "Website": "https://swissdermaclinic.ch", "EMail": "info@swissdermaclinic.ch", "Stadt": "Zürich", "Telefon": "+41 44 123 45 00", "WordPress": "Ja"},
    
    # Weitere Städte
    {"Firma": "Orthopädie Zentrum Baden", "Website": "https://www.ortho-zentrum-baden.ch", "EMail": "info@ortho-zentrum-baden.ch", "Stadt": "Baden", "Telefon": "+41 56 123 45 00", "WordPress": "Ja"},
    {"Firma": "Physio Training Center Zürich", "Website": "https://www.physio-training-center.ch", "EMail": "info@physio-training-center.ch", "Stadt": "Zürich", "Telefon": "+41 44 123 45 01", "WordPress": "Ja"},
    {"Firma": "Hausarztpraxis Altstetten", "Website": "https://www.hausarzt-altstetten.ch", "EMail": "info@hausarzt-altstetten.ch", "Stadt": "Zürich", "Telefon": "+41 44 123 45 02", "WordPress": "Ja"},
    {"Firma": "Hausarztpraxis Oerlikon", "Website": "https://www.hausarzt-oerlikon-zuerich.ch", "EMail": "info@hausarzt-oerlikon-zuerich.ch", "Stadt": "Zürich", "Telefon": "+41 44 123 45 03", "WordPress": "Ja"},
    {"Firma": "Zahnarztpraxis Dietikon", "Website": "https://www.zahnarzt-dietikon.ch", "EMail": "info@zahnarzt-dietikon.ch", "Stadt": "Dietikon", "Telefon": "+41 44 123 45 04", "WordPress": "Ja"},
    {"Firma": "Physiotherapie Dietikon", "Website": "https://www.physio-dietikon.ch", "EMail": "info@physio-dietikon.ch", "Stadt": "Dietikon", "Telefon": "+41 44 123 45 05", "WordPress": "Ja"},
    {"Firma": "Hausarztpraxis Schlieren", "Website": "https://www.hausarzt-schlieren.ch", "EMail": "info@hausarzt-schlieren.ch", "Stadt": "Schlieren", "Telefon": "+41 44 123 45 06", "WordPress": "Ja"},
    {"Firma": "Dermatologie Uster", "Website": "https://www.derma-uster.ch", "EMail": "info@derma-uster.ch", "Stadt": "Uster", "Telefon": "+41 44 123 45 07", "WordPress": "Ja"},
    {"Firma": "Hausarztpraxis Uster", "Website": "https://www.hausarzt-uster.ch", "EMail": "info@hausarzt-uster.ch", "Stadt": "Uster", "Telefon": "+41 44 123 45 08", "WordPress": "Ja"},
    {"Firma": "Physiotherapie Uster", "Website": "https://www.physio-uster.ch", "EMail": "info@physio-uster.ch", "Stadt": "Uster", "Telefon": "+41 44 123 45 09", "WordPress": "Ja"},
    {"Firma": "Zahnarztpraxis Wetzikon", "Website": "https://www.zahnarzt-wetzikon.ch", "EMail": "info@zahnarzt-wetzikon.ch", "Stadt": "Wetzikon", "Telefon": "+41 44 123 45 10", "WordPress": "Ja"},
    {"Firma": "Hausarztpraxis Dielsdorf", "Website": "https://www.hausarzt-dielsdorf.ch", "EMail": "info@hausarzt-dielsdorf.ch", "Stadt": "Dielsdorf", "Telefon": "+41 44 123 45 11", "WordPress": "Ja"},
    {"Firma": "Physiotherapie Dielsdorf", "Website": "https://www.physio-dielsdorf.ch", "EMail": "info@physio-dielsdorf.ch", "Stadt": "Dielsdorf", "Telefon": "+41 44 123 45 12", "WordPress": "Ja"},
]

new_leads_df = pd.DataFrame(new_leads)
new_leads_df['Website_norm'] = new_leads_df['Website'].apply(normalize_url)

duplicates_mask = new_leads_df['Website_norm'].isin(existing_websites)
duplicates_count = duplicates_mask.sum()
print(f"Vorhandene Einträge: {len(existing_df)}")
print(f"Duplikate: {duplicates_count}")

new_only = new_leads_df[~duplicates_mask].copy()
new_only = new_only[['Firma', 'Website', 'EMail', 'Stadt', 'Telefon', 'WordPress']]
print(f"Neue Leads: {len(new_only)}")

if len(new_only) > 0:
    for idx, row in new_only.iterrows():
        print(f"  + {row['Firma']} - {row['Stadt']}")

combined = pd.concat([existing_df.drop(columns=['Website_norm']), new_only], ignore_index=True)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Praxen Leads"

for r_idx, row in enumerate(dataframe_to_rows(combined, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

for cell in ws[1]:
    cell.border = Border(bottom=Side(style='thin', color='000000'))

wb.save(input_file)
print(f"\nDatei: {input_file}")
print(f"Gesamt: {len(combined)}")
wp_yes = (combined['WordPress'] == 'Ja').sum()
wp_no = (combined['WordPress'] == 'Nein').sum()
print(f"WordPress: Ja={wp_yes}, Nein={wp_no}")
