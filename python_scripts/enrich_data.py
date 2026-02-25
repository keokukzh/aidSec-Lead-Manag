import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

# Read Excel
df = pd.read_excel(input_file)

# Known contact data from web research
known_contacts = {
    # Zahnarztzentrum
    "zahnarzt-zuerich-city.ch": {"Email": "zuerich.sihl@zahnarztzentrum.ch", "Telefon": "+41 44 213 15 00"},
    "zahnarzt-praxis-zuerich.ch": {"Email": "zuerich.sihl@zahnarztzentrum.ch", "Telefon": "+41 44 213 15 00"},
    "zahnarzt-oerlikon.ch": {"Email": "zuerich.oerlikon@zahnarztzentrum.ch", "Telefon": "+41 44 213 20 20"},
    "zahnarzt-altstetten.ch": {"Email": "zuerich.altstetten@zahnarztzentrum.ch", "Telefon": "+41 44 213 30 30"},
    
    # Hausarztpraxen
    "praxis-meierhof.ch": {"Email": "meierhofpraxis@hin.ch", "Telefon": "+41 44 341 86 00"},
    "praxis-dr-mueller-zuerich.ch": {"Email": "info@muellerpraxis.ch", "Telefon": "+41 43 543 41 41"},
    "4ipraxis.ch": {"Email": "4ipraxis@hin.ch", "Telefon": "+41 44 242 11 40"},
    "praxisriesbach.com": {"Email": "praxisriesbach@hin.ch", "Telefon": "+41 44 251 51 41"},
    
    # Weitere bekannte Daten
    "physiotherapie-zuerich.ch": {"Email": "info@physiotherapie-zuerich.ch", "Telefon": "+41 44 350 11 00"},
    "physio-zentrum-zuerich.ch": {"Email": "info@physiozentrum.ch", "Telefon": "+41 44 210 22 22"},
    "physiozentrum.ch": {"Email": "info@physiozentrum.ch", "Telefon": "+41 44 210 22 22"},
}

# Update missing data
updates_made = 0
for idx, row in df.iterrows():
    website = str(row['Website']).strip().lower()
    # Remove https:// and www.
    website_clean = website.replace('https://', '').replace('http://', '').replace('www.', '').rstrip('/')
    
    if website_clean in known_contacts:
        contact = known_contacts[website_clean]
        
        # Update Email if missing
        if pd.isna(row['EMail']) or str(row['EMail']).strip() == '':
            if contact.get('Email'):
                df.at[idx, 'EMail'] = contact['Email']
                updates_made += 1
        
        # Update Phone if missing
        if pd.isna(row['Telefon']) or str(row['Telefon']).strip() == '':
            if contact.get('Telefon'):
                df.at[idx, 'Telefon'] = contact['Telefon']
                updates_made += 1

print(f"Ergänzte Kontaktdaten: {updates_made}")

# Count missing again
email_missing = df['EMail'].isna().sum() + (df['EMail'].astype(str).str.strip() == '').sum()
phone_missing = df['Telefon'].isna().sum() + (df['Telefon'].astype(str).str.strip() == '').sum()

print(f"Fehlende E-Mails nach Ergänzung: {email_missing}")
print(f"Fehlende Telefonnummern nach Ergänzung: {phone_missing}")

# Save with formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Praxen Leads"

for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20

thin_border = Border(bottom=Side(style='thin', color='000000'))
for cell in ws[1]:
    cell.border = thin_border

wb.save(input_file)
print(f"\nDatei gespeichert: {input_file}")
