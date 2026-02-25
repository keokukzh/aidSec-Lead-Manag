import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

# Read existing data
existing_df = pd.read_excel(input_file)

# Normalize website for comparison - remove www. and trailing slashes
def normalize_url(url):
    url = str(url).strip().lower()
    url = url.rstrip('/')
    if url.startswith('https://www.'):
        url = url.replace('https://www.', 'https://')
    elif url.startswith('http://www.'):
        url = url.replace('http://www.', 'http://')
    return url

existing_df['Website_norm'] = existing_df['Website'].apply(normalize_url)
existing_websites = existing_df['Website_norm'].tolist()

# New leads from user
new_leads_raw = """Arzthaus Zürich City;https://www.arzthaus.ch;Unbekannt;zuerich@arzthaus.ch;Zürich;+41 44 800 39 00;Impressum & Datenschutz vorhanden
medelio Arztpraxis Zürich Altstetten;https://www.medelio.ch;Unbekannt;medelio@hin.ch;Zürich;+41 44 545 30 98;Datenschutzerklärung vorhanden
Praxis Dr. med. Christoph Schuppli;https://www.praxis-schuppli.ch;Unbekannt;info@praxis-schuppli.ch;Zürich;+41 44 211 37 00;Datenschutz vorhanden"""

from io import StringIO
new_leads_df = pd.read_csv(StringIO(new_leads_raw), sep=';', header=None, 
                           names=['Firma', 'Website', 'CMS', 'EMail', 'Stadt', 'Telefon', 'nDSG'])

new_leads_df = new_leads_df[['Firma', 'Website', 'EMail', 'Stadt', 'Telefon']]
new_leads_df['Website_norm'] = new_leads_df['Website'].apply(normalize_url)

# Check for duplicates
duplicates_mask = new_leads_df['Website_norm'].isin(existing_websites)

duplicates_count = duplicates_mask.sum()
print(f"Vorhandene Einträge: {len(existing_df)}")
print(f"Duplikate gefunden: {duplicates_count}")

# Show duplicates
if duplicates_count > 0:
    print("\nDuplikate (werden übersprungen):")
    for idx, row in new_leads_df[duplicates_mask].iterrows():
        print(f"  - {row['Firma']} ({row['Website']})")

# Filter only new leads
new_only = new_leads_df[~duplicates_mask].copy()
new_only = new_only[['Firma', 'Website', 'EMail', 'Stadt', 'Telefon']]

print(f"Neue Leads hinzugefügt: {len(new_only)}")

# Combine
combined = pd.concat([existing_df.drop(columns=['Website_norm']), new_only], ignore_index=True)

# Save with formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Praxen Leads"

for r_idx, row in enumerate(dataframe_to_rows(combined, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20

thin_border = Border(bottom=Side(style='thin', color='000000'))
for cell in ws[1]:
    cell.border = thin_border

wb.save(input_file)
print(f"\nDatei aktualisiert: {input_file}")
print(f"Gesamtanzahl Einträge: {len(combined)}")
