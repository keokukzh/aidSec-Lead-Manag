import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

new_leads = [
    {'Firma': 'Hueberli Lawyers AG', 'Website': 'hueberli.com', 'EMail': 'contact@hueberli.com', 'Stadt': 'Wattwil', 'Telefon': '+41 71 988 30 00', 'WordPress': 'Ja'},
    {'Firma': 'LAUX LAWYERS AG', 'Website': 'lauxlawyers.ch', 'EMail': 'info@lauxlawyers.ch', 'Stadt': 'Zürich', 'Telefon': '+41 44 880 24 24', 'WordPress': 'Ja'},
    {'Firma': 'RA Marc Schmid GmbH', 'Website': 'zuerich-strafrecht.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '+41 79 105 80 83', 'WordPress': 'Ja'},
    {'Firma': 'Advokaturbüro Bütikofer', 'Website': 'advobue.ch', 'EMail': 'info@advobue.ch', 'Stadt': 'Gossau SG', 'Telefon': '+41 71 310 06 88', 'WordPress': 'Ja'},
    {'Firma': 'HütteLAW AG', 'Website': 'huettelaw.ch', 'EMail': 'office@huettelaw.ch', 'Stadt': 'Cham', 'Telefon': '+41 41 729 36 36', 'WordPress': 'Ja'},
    {'Firma': 'BernLex Juristen', 'Website': 'bernlex.ch', 'EMail': 'buta@bernlex.ch', 'Stadt': 'Bern', 'Telefon': '+41 31 310 91 91', 'WordPress': 'Ja'},
    {'Firma': 'BSG Partner', 'Website': 'bsgpartner.ch', 'EMail': '', 'Stadt': 'Steffisburg', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Zanetti Rechtsanwälte AG', 'Website': 'scheidungsanwaltzug.ch', 'EMail': 'zanetti@zanettilaw.ch', 'Stadt': 'Zug', 'Telefon': '+41 41 766 35 85', 'WordPress': 'Ja'},
    {'Firma': 'Inauen Moser Rechtsanwälte', 'Website': 'inauen-moser.ch', 'EMail': 'info@inauen-moser.ch', 'Stadt': 'Appenzell', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Advokatur & Notariat Ochsner', 'Website': 'aanot.ch', 'EMail': 'law@aanot.ch', 'Stadt': 'Bern', 'Telefon': '+41 31 312 10 00', 'WordPress': 'Ja'},
    {'Firma': 'KSCP Rechtsanwälte und Notare', 'Website': 'kscp-legal.ch', 'EMail': 'grenchen@kscp-legal.ch', 'Stadt': 'Grenchen', 'Telefon': '+41 32 654 99 10', 'WordPress': 'Ja'},
    {'Firma': 'Rhyner Rechtsanwälte Notare', 'Website': 'law-switzerland.ch', 'EMail': 'info@law-switzerland.ch', 'Stadt': 'Glarus', 'Telefon': '+41 55 645 37 37', 'WordPress': 'Ja'},
    {'Firma': 'Pfister & Partner Rechtsanwälte AG', 'Website': 'pfister-rechtsanwaelte.ch', 'EMail': 'info@pfister-rechtsanwaelte.ch', 'Stadt': 'Pfäffikon', 'Telefon': '+41 55 415 80 80', 'WordPress': 'Ja'},
    {'Firma': 'Advokatur Klug', 'Website': 'advokatur-klug.ch', 'EMail': 'info@advokatur-klug.ch', 'Stadt': 'St. Gallen', 'Telefon': '+41 71 520 61 21', 'WordPress': 'Ja'},
    {'Firma': 'Büchel von Rohr Rechtsanwälte', 'Website': 'kanzlei-uster.ch', 'EMail': 'sekretariat@kanzlei-uster.ch', 'Stadt': 'Uster', 'Telefon': '+41 44 512 25 00', 'WordPress': 'Ja'},
    {'Firma': 'Advokatur Stegmann AG', 'Website': 'advokaturstegmann.ch', 'EMail': '', 'Stadt': 'Biel', 'Telefon': '+41 32 525 49 52', 'WordPress': 'Ja'},
    {'Firma': 'Landmann & Partner', 'Website': 'landmann.ch', 'EMail': 'info@landmann.ch', 'Stadt': 'Zürich', 'Telefon': '+41 44 361 61 65', 'WordPress': 'Ja'},
    {'Firma': 'Kanzlei3 Advokatur und Notariat', 'Website': 'kanzlei3.ch', 'EMail': 'kanzlei@kanzlei3.ch', 'Stadt': 'Brig', 'Telefon': '+41 27 922 11 44', 'WordPress': 'Ja'},
    {'Firma': 'Stössel Schweizer Partner', 'Website': 'kanzlei-winterthur.ch', 'EMail': '', 'Stadt': 'Winterthur', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Schiller Rechtsanwälte AG', 'Website': 'schillerlegal.ch', 'EMail': 'office@schillerlegal.ch', 'Stadt': 'Winterthur', 'Telefon': '+41 52 269 16 16', 'WordPress': 'Ja'},
    {'Firma': 'Grütter Rechtsanwälte AG', 'Website': 'gruetter.ch', 'EMail': '', 'Stadt': 'Solothurn', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Romero & Ziegler Meier Jucker', 'Website': 'romeroziegler.ch', 'EMail': 'sekretariat@romeroziegler.ch', 'Stadt': 'Zürich', 'Telefon': '+41 44 206 30 10', 'WordPress': 'Ja'},
    {'Firma': 'HWN Rechtsanwälte', 'Website': 'anwaltskanzlei-erbrecht.ch', 'EMail': 'info@anwaltskanzlei-erbrecht.ch', 'Stadt': 'Zürich', 'Telefon': '+41 44 535 55 50', 'WordPress': 'Ja'},
    {'Firma': 'Eberhart Anwaltskanzlei AG', 'Website': 'eberhart-law.ch', 'EMail': 'legal@eberhart-law.ch', 'Stadt': 'Bern', 'Telefon': '+41 31 330 81 11', 'WordPress': 'Ja'},
]

def normalize_url(url):
    url = str(url).strip().lower().rstrip('/')
    url = url.replace('https://', '').replace('http://', '').replace('www.', '')
    return url

wb = openpyxl.load_workbook(input_file)
ws = wb['Anwalts Kanzleien']

existing_urls = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    val = str(row[0].value or '').strip()
    if val:
        existing_urls.add(normalize_url(val))

print(f"Bestehende Einträge: {len(existing_urls)}")

thin_border = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0')
)
data_font = Font(name='Calibri', size=11, color='333333')
data_alignment = Alignment(horizontal='left', vertical='center')

added = 0
skipped = 0

for lead in new_leads:
    norm_url = normalize_url(lead['Website'])
    if norm_url in existing_urls:
        print(f"  SKIP (Duplikat): {lead['Website']}")
        skipped += 1
        continue

    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=lead['Firma'])
    ws.cell(row=next_row, column=2, value=lead['Website'])
    ws.cell(row=next_row, column=3, value=lead['EMail'])
    ws.cell(row=next_row, column=4, value=lead['Stadt'])
    ws.cell(row=next_row, column=5, value=lead['Telefon'])
    ws.cell(row=next_row, column=6, value=lead['WordPress'])

    for col in range(1, 7):
        cell = ws.cell(row=next_row, column=col)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

    existing_urls.add(norm_url)
    added += 1
    print(f"  ADDED: {lead['Firma']} ({lead['Website']})")

wb.save(input_file)
print(f"\nErgebnis: {added} hinzugefügt, {skipped} übersprungen")
print(f"Gesamt 'Anwalts Kanzleien': {ws.max_row - 1} Einträge")
