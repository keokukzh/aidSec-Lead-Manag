import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Existing file
input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

# Read existing data
existing_df = pd.read_excel(input_file)
print(f"Vorhandene Einträge: {len(existing_df)}")

# New leads data from user
new_leads_data = """email,firstname,lastname,company,website,phone,city,hs_lead_status,ndsg_relevant,security_audit_note,outreach_status
,,,Zahnarzt Zürich City,https://www.zahnarzt-zuerich-city.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Praxis Zürich,https://www.zahnarzt-praxis-zuerich.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Oerlikon,https://www.zahnarzt-oerlikon.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Altstetten,https://www.zahnarzt-altstetten.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Seefeld,https://www.zahnarzt-seefeld.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Langstrasse,https://www.zahnarzt-langstrasse.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Wiedikon,https://www.zahnarzt-wiedikon.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Albisrieden,https://www.zahnarzt-albisrieden.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Affoltern,https://www.zahnarzt-affoltern.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Hirslanden,https://www.zahnarzt-hirslanden.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Hausarzt Praxis Zürich,https://www.hausarzt-praxis-zuerich.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Praxis Dr. Müller Zürich,https://www.praxis-dr-mueller-zuerich.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Allgemeinarzt Zürich,https://www.allgemeinarzt-zuerich.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Hausarzt Oerlikon,https://www.hausarzt-oerlikon.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Hausarzt Seefeld,https://www.hausarzt-seefeld.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Hausarzt Langstrasse,https://www.hausarzt-langstrasse.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Praxis am Heimplatz,https://www.praxis-am-heimplatz.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Praxis Stadelhofen,https://www.praxis-stadelhofen.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Hausarzt Wiedikon,https://www.hausarzt-wiedikon.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Praxis Limmatplatz,https://www.praxis-limmatplatz.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Physiotherapie Zürich,https://www.physiotherapie-zuerich.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Physio Zentrum Zürich,https://www.physio-zentrum-zuerich.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Physiotherapie Oerlikon,https://www.physiotherapie-oerlikon.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Physio Praxis Zürich,https://www.physio-praxis-zuerich.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Physio Fit Zürich,https://www.physio-fit-zuerich.ch,,Zürich,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Bern,https://www.zahnarzt-bern.ch,,Bern,NEW,JA,F,NOT_STARTED
,,,Zahnaerzte Bern,https://www.zahnaerzte-bern.ch,,Bern,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Praxis Bern,https://www.zahnarzt-praxis-bern.ch,,Bern,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Lorraine,https://www.zahnarzt-lorraine.ch,,Bern,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Breitenrain,https://www.zahnarzt-breitenrain.ch,,Bern,NEW,JA,F,NOT_STARTED
,,,Hausarzt Bern,https://www.hausarzt-bern.ch,,Bern,NEW,JA,F,NOT_STARTED
,,,Hausarztpraxis Bern,https://www.hausarztpraxis-bern.ch,,Bern,NEW,JA,F,NOT_STARTED
,,,Praxis Dr. Bern,https://www.praxis-dr-bern.ch,,Bern,NEW,JA,F,NOT_STARTED
,,,Allgemeinpraxis Bern,https://www.allgemeinpraxis-bern.ch,,Bern,NEW,JA,F,NOT_STARTED
,,,Praxis Lorraine Bern,https://www.praxis-lorraine-bern.ch,,Bern,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Luzern,https://www.zahnarzt-luzern.ch,,Luzern,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Praxis Luzern,https://www.zahnarzt-praxis-luzern.ch,,Luzern,NEW,JA,F,NOT_STARTED
,,,Hausarzt Luzern,https://www.hausarzt-luzern.ch,,Luzern,NEW,JA,F,NOT_STARTED
,,,Hausarztpraxis Luzern,https://www.hausarztpraxis-luzern.ch,,Luzern,NEW,JA,F,NOT_STARTED
,,,Praxis Zentrum Luzern,https://www.praxis-zentrum-luzern.ch,,Luzern,NEW,JA,F,NOT_STARTED
,,,Zahnarzt Basel,https://www.zahnarzt-basel.ch,,Basel,NEW,JA,F,NOT_STARTED
,,,Zahnarztpraxis Basel,https://www.zahnarztpraxis-basel.ch,,Basel,NEW,JA,F,NOT_STARTED
,,,Hausarzt Basel,https://www.hausarzt-basel.ch,,Basel,NEW,JA,F,NOT_STARTED
,,,Hausarztpraxis Basel,https://www.hausarztpraxis-basel.ch,,Basel,NEW,JA,F,NOT_STARTED
,,,Praxis Basel,https://www.praxis-basel.ch,,Basel,NEW,JA,F,NOT_STARTED
,,,Zahnarzt St. Gallen,https://www.zahnarzt-stgallen.ch,,St. Gallen,NEW,JA,F,NOT_STARTED
,,,Zahnarztpraxis St. Gallen,https://www.zahnarztpraxis-stgallen.ch,,St. Gallen,NEW,JA,F,NOT_STARTED
,,,Hausarzt St. Gallen,https://www.hausarzt-stgallen.ch,,St. Gallen,NEW,JA,F,NOT_STARTED
,,,Hausarztpraxis St. Gallen,https://www.hausarztpraxis-stgallen.ch,,St. Gallen,NEW,JA,F,NOT_STARTED
,,,Praxis St. Gallen,https://www.praxis-stgallen.ch,,St. Gallen,NEW,JA,F,NOT_STARTED"""

from io import StringIO
new_leads_df = pd.read_csv(StringIO(new_leads_data))

# Map to our schema: Firma, Website, EMail, Stadt, Telefon
new_leads_df = new_leads_df[['company', 'website', 'email', 'city', 'phone']].copy()
new_leads_df.columns = ['Firma', 'Website', 'EMail', 'Stadt', 'Telefon']

# Get existing websites
existing_websites = existing_df['Website'].str.strip().str.lower().tolist()
print(f"Vorhandene Websites: {len(existing_websites)}")

# Check for duplicates
new_leads_df['Website_clean'] = new_leads_df['Website'].str.strip().str.lower()
duplicates_mask = new_leads_df['Website_clean'].isin(existing_websites)

duplicates_count = duplicates_mask.sum()
print(f"Duplikate gefunden: {duplicates_count}")

# Filter only new leads
new_only = new_leads_df[~duplicates_mask].copy()
new_only = new_only.drop(columns=['Website_clean'])

print(f"Neue Leads hinzugefügt: {len(new_only)}")

# Combine with existing
combined = pd.concat([existing_df, new_only], ignore_index=True)

# Save to Excel with formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Praxen Leads"

for r_idx, row in enumerate(dataframe_to_rows(combined, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20

thin_border = Border(bottom=Side(style='thin', color='000000'))
for cell in ws[1]:
    cell.border = thin_border

wb.save(input_file)
print(f"\nDatei aktualisiert: {input_file}")
print(f"Gesamtanzahl Einträge: {len(combined)}")
