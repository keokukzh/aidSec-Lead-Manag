import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

df = pd.read_excel(input_file, sheet_name="Anwalts Kanzleien")
for col in ['Firma', 'Website', 'EMail', 'Stadt', 'Telefon', 'WordPress']:
    df[col] = df[col].astype(str).replace('nan', '')
print(f"Ausgangslage: {len(df)} Einträge")

def normalize_url(url):
    url = str(url).strip().lower().rstrip('/')
    url = url.replace('https://', '').replace('http://', '').replace('www.', '')
    return url

# --- PHASE 1: Remove duplicates & invalid entries ---
df['url_norm'] = df['Website'].apply(normalize_url)

urls_to_remove = [
    'b-legal.ch',           # #23 duplicate of #4 (www.b-legal.ch)
    'raewel-advokatur.ch',  # #41 duplicate of #1 (www.raewel-advokatur.ch)
]

# rechtsanwalt-wil.ch (#21) redirects to same firm as rechtsanwalt-zuerich.ch (#32)
urls_to_remove.append('rechtsanwalt-wil.ch')

# law-bern.ch (#20) is same firm as advokatur-bern.ch (#34)
urls_to_remove.append('law-bern.ch')

# anwaltskanzlei-zug.ch (#37): PARKED DOMAIN - not a law firm
urls_to_remove.append('anwaltskanzlei-zug.ch')

# muellerpartner.ch (#50): redirects to poster gallery (Artifiche)
urls_to_remove.append('muellerpartner.ch')

mask_keep = ~df['url_norm'].isin(urls_to_remove)

# Also remove exact duplicates by url_norm (keeping first)
df_clean = df[mask_keep].drop_duplicates(subset=['url_norm'], keep='first').copy()

removed_count = len(df) - len(df_clean)
print(f"Entfernt: {removed_count} Einträge (Duplikate + ungültige)")
print(f"Verbleibend: {len(df_clean)} Einträge")

# --- PHASE 2: Correct errors and enrich data ---
known_data = {
    'raewel-advokatur.ch': {
        'Firma': 'RÄWEL ADVOKATUR',
        'Telefon': '+41 44 271 01 71',
        'EMail': 'info@raewel-advokatur.ch',
        'Stadt': 'Zürich',
    },
    'kokotek.ch': {
        'Telefon': '+41 44 251 60 80',
        'EMail': 'info@kokotek.ch',
        'Stadt': 'Zürich',
    },
    'schwenninger.ch': {
        'Telefon': '+41 44 820 41 42',
        'EMail': 'kanzlei@schwenninger.ch',
        'Stadt': 'Rüti ZH',
    },
    'b-legal.ch': {
        'Telefon': '+41 44 266 20 66',
        'EMail': 'info@b-legal.ch',
        'Stadt': 'Zürich',
    },
    'anwaltskanzlei-wiki.ch': {
        'Telefon': '+41 44 383 33 55',
        'EMail': 'mail@anwaltskanzlei-wiki.ch',
        'Stadt': 'Herrliberg',
    },
    'landmann.ch': {
        'Telefon': '+41 44 361 61 65',
        'EMail': 'info@landmann.ch',
        'Stadt': 'Zürich',
    },
    'lp-zurich.ch': {
        'Telefon': '+41 44 500 90 45',
        'EMail': 'kanzlei@lp-zurich.ch',
        'Stadt': 'Zürich',
    },
    'boesch-anwaelte.ch': {
        'Stadt': 'Zürich',
    },
    'galik.ch': {
        'Stadt': 'Zürich',
    },
    'valentinspahr.ch': {
        'Telefon': '+41 56 470 90 00',
        'EMail': 'spahr@spahrlegal.ch',
        'Stadt': 'Baden-Dättwil',
    },
    'dieadvokatur.ch': {
        'Telefon': '+41 41 227 58 58',
        'EMail': 'info@dieadvokatur.ch',
        'Stadt': 'Luzern',
    },
    'advokaturteam.ch': {
        'Telefon': '+41 61 273 70 70',
        'EMail': 'info@advokaturteam.ch',
        'Stadt': 'Basel',  # CORRECTION: was "Bern", actually Basel
    },
    'advobasel.ch': {
        'Telefon': '+41 61 264 94 94',
        'EMail': 'info@advobasel.ch',
        'Stadt': 'Basel',
    },
    'advokatur-lachen.ch': {
        'Telefon': '+41 55 451 54 03',
        'Stadt': 'Lachen',
    },
    'adrianhaas.ch': {
        'Firma': 'Adrian Haas',
        'Telefon': '+41 79 717 24 24',
        'EMail': 'adrianhaas@gmx.ch',
        'Stadt': 'Bern',
    },
    'advokatur-trias.ch': {
        'Firma': 'Advokatur-Trias AG',
        'Telefon': '+41 62 393 03 03',
        'EMail': 'info@advokatur-trias.ch',
        'Stadt': 'Aarau',
    },
    'good-stgallen.ch': {
        'Telefon': '+41 58 510 88 99',
        'Stadt': 'St. Gallen',
    },
    'advokatur-klug.ch': {
        'Firma': 'Advokatur & Notariat Klug',
        'Telefon': '+41 71 520 61 21',
        'EMail': 'info@advokatur-klug.ch',
        'Stadt': 'St. Gallen',  # CORRECTION: was "Zürich", actually St. Gallen
    },
    'organhaftung-schweiz.ch': {
        'Stadt': 'Zürich',
    },
    'kanzlei-winterthur.ch': {
        'Firma': 'Stössel Schweizer Partner',
        'Telefon': '+41 52 223 19 19',
        'EMail': 'info@kanzlei-winterthur.ch',
        'Stadt': 'Winterthur',
    },
    'steigerlegal.ch': {
        'Telefon': '+41 44 533 13 60',
        'EMail': 'martin.steiger@steigerlegal.ch',
        'Stadt': 'Zürich',
    },
    'steinlex.ch': {
        'Stadt': 'Zürich',
    },
    'ra-kost.ch': {
        'Telefon': '+41 41 440 33 43',
        'EMail': 'info@ra-kost.ch',
        'Stadt': 'Ebikon',
    },
    'bmlaw.ch': {
        'Firma': 'Baumgartner Mächler Rechtsanwälte AG',
        'Telefon': '+41 44 215 44 77',
        'EMail': 'office@bmlaw.ch',
        'Stadt': 'Zürich',
    },
    'advokatur-west.ch': {
        'Telefon': '+41 44 405 19 19',
        'Stadt': 'Zürich',
    },
    'hirschrecht.ch': {
        'Firma': 'Kanzlei Hirschrecht',
        'Telefon': '+41 43 343 12 93',
        'EMail': 'info@hirschrecht.ch',
        'Stadt': 'Zürich',
    },
    'schmidlaw.ch': {
        'Firma': 'Schmid Rechtsanwälte',
        'Telefon': '+41 44 220 10 40',
        'EMail': 'info@schmidlaw.ch',
        'Stadt': 'Zürich',
    },
    'rsplaw.ch': {
        'Telefon': '+41 44 217 70 10',
        'EMail': 'info@rsplaw.ch',
        'Stadt': 'Zürich',
    },
    'rechtsanwalt-zuerich.ch': {
        'Firma': 'Heller Rechtsanwalts AG',
        'Telefon': '+41 43 499 77 33',
        'EMail': 'heller@rechtsanwalt-zuerich.ch',
        'Stadt': 'Zürich',
    },
    'rechtsanwalt-baden.ch': {
        'Firma': 'Stephani + Partner',
        'Telefon': '+41 56 483 50 50',
        'EMail': 'info@stephani-partner.ch',
        'Stadt': 'Baden-Dättwil',
    },
    'advokatur-bern.ch': {
        'Firma': 'Brechbühler Rechtsanwälte',
        'Telefon': '+41 31 356 16 16',
        'EMail': 'mail@advokatur-brechbuehler.ch',
        'Stadt': 'Bern',
    },
    'haslerfrech.ch': {
        'Firma': 'Hasler & Frech',
        'Telefon': '+41 32 621 20 20',
        'EMail': 'info@haslerfrech.ch',
        'Stadt': 'Solothurn',
    },
    'forumlaw.ch': {
        'Telefon': '+41 44 350 10 00',
        'EMail': 'info@forumlaw.ch',
        'Stadt': 'Zürich',
    },
    'anwalt-aargau.ch': {
        'Firma': 'Anwalt Aargau',
        'Stadt': 'Aargau',
    },
    'gegenschatzpartner.ch': {
        'Firma': 'Gegenschatz Partner AG',
        'Telefon': '+41 44 280 28 00',
        'Stadt': 'Zürich',
    },
    'fischer-partner.ch': {
        'Telefon': '+41 71 987 73 00',
        'EMail': 'info@fischer-partner.ch',
        'Stadt': 'Wattwil',
    },
    'berger-recht.ch': {
        'Firma': 'Anwaltskanzlei Kurt Berger',
        'Telefon': '+41 44 316 66 51',
        'EMail': 'berger@berger-recht.ch',
        'Stadt': 'Zürich',
    },
    'anwalt-frauenfeld.ch': {
        'Firma': 'Anwalt Frauenfeld',
        'Stadt': 'Frauenfeld',
    },
    'rechtsberatung-zuerich.ch': {
        'Firma': 'Rechtsberatung Zürich',
        'Stadt': 'Zürich',
    },
    'graflaw.ch': {
        'Firma': 'GrafLaw - Simon M. Graf',
        'EMail': 'office@graflaw.ch',
    },
    'advokatur-zuerich.ch': {
        'Firma': 'Fischer Rechtsanwälte',
        'Stadt': 'Zürich',
    },
    'n-law.ch': {
        'Firma': 'Nyffenegger Rechtsanwälte',
        'Telefon': '+41 43 888 66 00',
        'EMail': 'info@n-law.ch',
        'Stadt': 'Zürich',
    },
    'lenzlaw.ch': {
        'Firma': 'CL - Christian Lenz',
    },
    'becklaw.ch': {
        'Firma': 'Beck Law',
    },
}

updates_count = 0
for idx, row in df_clean.iterrows():
    url_key = row['url_norm']
    if url_key in known_data:
        data = known_data[url_key]
        for field, value in data.items():
            current = str(row.get(field, '')).strip()
            if current == 'nan':
                current = ''
            if field in ('Firma', 'Stadt'):
                if current == '' or current != value:
                    df_clean.at[idx, field] = value
                    updates_count += 1
            elif field in ('Telefon', 'EMail'):
                if current == '':
                    df_clean.at[idx, field] = value
                    updates_count += 1

print(f"Aktualisiert: {updates_count} Felder")

# --- PHASE 3: Clean up formatting ---
def title_case_city(city):
    if pd.isna(city) or str(city).strip() == '' or str(city).strip() == 'nan':
        return ''
    city = str(city).strip()
    if city == city.lower():
        return city.title()
    return city

df_clean['Stadt'] = df_clean['Stadt'].apply(title_case_city)

def clean_firma(firma):
    if pd.isna(firma) or str(firma).strip() == '' or str(firma).strip() == 'nan':
        return ''
    return str(firma).strip()

df_clean['Firma'] = df_clean['Firma'].apply(clean_firma)

def clean_email(email):
    if pd.isna(email) or str(email).strip() == '' or str(email).strip() == 'nan':
        return ''
    return str(email).strip()

df_clean['EMail'] = df_clean['EMail'].apply(clean_email)

def clean_phone(phone):
    if pd.isna(phone) or str(phone).strip() == '' or str(phone).strip() == 'nan':
        return ''
    return str(phone).strip()

df_clean['Telefon'] = df_clean['Telefon'].apply(clean_phone)

def clean_website(url):
    if pd.isna(url) or str(url).strip() == '':
        return ''
    url = str(url).strip().rstrip('/')
    if not url.startswith('http'):
        url = 'https://' + url
    return url

df_clean['Website'] = df_clean['Website'].apply(clean_website)

df_clean['WordPress'] = 'Ja'

df_final = df_clean[['Firma', 'Website', 'EMail', 'Stadt', 'Telefon', 'WordPress']].copy()
df_final = df_final.reset_index(drop=True)

print(f"\nFinale Tabelle: {len(df_final)} Einträge")
missing_email = df_final[df_final['EMail'] == ''].shape[0]
missing_phone = df_final[df_final['Telefon'] == ''].shape[0]
missing_city = df_final[df_final['Stadt'] == ''].shape[0]
missing_firma = df_final[df_final['Firma'] == ''].shape[0]
print(f"Fehlende E-Mails: {missing_email}")
print(f"Fehlende Telefone: {missing_phone}")
print(f"Fehlende Städte: {missing_city}")
print(f"Fehlende Firmennamen: {missing_firma}")

# --- PHASE 4: Write back to Excel ---
wb = openpyxl.load_workbook(input_file)

if "Anwalts Kanzleien" in wb.sheetnames:
    del wb["Anwalts Kanzleien"]

ws = wb.create_sheet("Anwalts Kanzleien")

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
cell_alignment = Alignment(vertical="center")
thin_border = Border(bottom=Side(style='thin', color='D9D9D9'))
header_border = Border(bottom=Side(style='medium', color='2F5496'))

for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border
        else:
            cell.alignment = cell_alignment
            cell.border = thin_border
            if c_idx == 2 and value:
                cell.font = Font(color="0563C1", underline="single")

ws.column_dimensions['A'].width = 38
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

ws.auto_filter.ref = f"A1:F{len(df_final) + 1}"
ws.freeze_panes = "A2"

wb.save(input_file)
print(f"\nDatei gespeichert: {input_file}")

print("\n--- Entfernte Einträge ---")
print("- Badertscher Rechtsanwälte (b-legal.ch) -> Duplikat")
print("- Raewel Advokatur (raewel-advokatur.ch) -> Duplikat")
print("- Rechtsanwalt Wil (rechtsanwalt-wil.ch) -> Duplikat von Heller Rechtsanwalts AG")
print("- Law Bern (law-bern.ch) -> Duplikat von advokatur-bern.ch")
print("- anwaltskanzlei-zug.ch -> Geparkte Domain (keine Kanzlei)")
print("- Müller Partner (muellerpartner.ch) -> Leitet auf Plakatgalerie weiter")

print("\n--- Korrekturen ---")
print("- Advokaturteam: Stadt 'Bern' -> 'Basel' (tatsächlicher Standort)")
print("- Advokatur Klug: Stadt 'Zürich' -> 'St. Gallen' (tatsächlicher Standort)")
print("- Hasler & Frech: Stadt 'solothurn' -> 'Solothurn'")
print("- Diverse fehlende Firmennamen ergänzt")
print("- Diverse fehlende E-Mails, Telefonnummern und Städte ergänzt")
