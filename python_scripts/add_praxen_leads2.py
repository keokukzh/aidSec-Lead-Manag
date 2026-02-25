import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

new_leads = [
    # Augenarzt Bern
    {'Firma': 'Berner Augenklinik', 'Website': 'augenklinik-bern.ch', 'EMail': '', 'Stadt': 'Bern', 'Telefon': '+41 31 311 12 22', 'WordPress': 'Ja'},
    # Augenarzt Luzern
    {'Firma': 'AugenärzteZentrum Luzern', 'Website': 'augenaerzte-luzern.ch', 'EMail': '', 'Stadt': 'Luzern', 'Telefon': '+41 41 342 22 22', 'WordPress': 'Ja'},
    # Augenarzt Verbund
    {'Firma': 'Gutblick', 'Website': 'gutblick.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '', 'WordPress': 'Ja'},
    # Augenarzt St. Gallen
    {'Firma': 'Augenarzt Dr. Graemiger', 'Website': 'augenarzt-drgraemiger.ch', 'EMail': '', 'Stadt': 'St. Gallen', 'Telefon': '+41 71 222 88 66', 'WordPress': 'Ja'},
    # Augenarzt Winterthur
    {'Firma': 'Augenärzte Zentrum Winterthur', 'Website': 'augentagesklinik-winterthur.ch', 'EMail': '', 'Stadt': 'Winterthur', 'Telefon': '+41 52 242 38 38', 'WordPress': 'Ja'},
    # Dermato Basel
    {'Firma': 'Dermatologie am Rhein', 'Website': 'dermatologie-am-rhein.ch', 'EMail': '', 'Stadt': 'Basel', 'Telefon': '', 'WordPress': 'Ja'},
    # Dermato Bern
    {'Firma': 'aareSkin Dermatologie', 'Website': 'hautarzt-baerenplatz-bern.ch', 'EMail': '', 'Stadt': 'Bern', 'Telefon': '+41 31 331 80 80', 'WordPress': 'Ja'},
    # Ortho St. Gallen
    {'Firma': 'Orthopädie St. Gallen AG', 'Website': 'ortho-sg.ch', 'EMail': '', 'Stadt': 'St. Gallen', 'Telefon': '+41 71 228 88 99', 'WordPress': 'Ja'},
    # Zahnarzt Zug
    {'Firma': 'Zahnarztpraxis Dr. Toth', 'Website': 'zahnarzt-zug.net', 'EMail': 'info@zahnarzt-zug.net', 'Stadt': 'Zug', 'Telefon': '+41 41 711 26 02', 'WordPress': 'Ja'},
    # Hausarzt Olten
    {'Firma': 'RingPraxis Olten', 'Website': 'ringpraxis-olten.ch', 'EMail': '', 'Stadt': 'Olten', 'Telefon': '+41 62 212 38 36', 'WordPress': 'Ja'},
    {'Firma': 'Hausarztzentrum Olten', 'Website': 'hausarztzentrum-olten.ch', 'EMail': '', 'Stadt': 'Olten', 'Telefon': '+41 62 212 36 46', 'WordPress': 'Ja'},
]

def normalize_url(url):
    url = str(url).strip().lower().rstrip('/')
    url = url.replace('https://', '').replace('http://', '').replace('www.', '')
    return url

wb = openpyxl.load_workbook(input_file)
ws = wb['Praxen Leads']
existing_urls = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    val = str(row[0].value or '').strip()
    if val: existing_urls.add(normalize_url(val))
print(f"Bestehende Eintraege: {ws.max_row - 1}")

thin_border = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'),
                     top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))
data_font = Font(name='Calibri', size=11, color='333333')
data_alignment = Alignment(horizontal='left', vertical='center')
added = skipped = 0
for lead in new_leads:
    norm_url = normalize_url(lead['Website'])
    if norm_url in existing_urls:
        print(f"  SKIP: {lead['Website']}"); skipped += 1; continue
    next_row = ws.max_row + 1
    for i, key in enumerate(['Firma','Website','EMail','Stadt','Telefon','WordPress'], 1):
        cell = ws.cell(row=next_row, column=i, value=lead[key])
        cell.font = data_font; cell.alignment = data_alignment; cell.border = thin_border
    existing_urls.add(norm_url); added += 1
    print(f"  ADDED: {lead['Firma']} ({lead['Website']})")
wb.save(input_file)
print(f"\nErgebnis: {added} hinzugefuegt, {skipped} uebersprungen")
print(f"Gesamt 'Praxen Leads': {ws.max_row - 1} Eintraege")
