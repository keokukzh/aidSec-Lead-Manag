import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

new_leads = [
    # Ortho Bern
    {'Firma': 'Orthoplus Bern', 'Website': 'orthoplus-bern.ch', 'EMail': '', 'Stadt': 'Bern', 'Telefon': '+41 31 996 90 90', 'WordPress': 'Ja'},
    # Dermato Zürich Oerlikon
    {'Firma': 'ST-MED Zürich Oerlikon', 'Website': 'st-med.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '+41 44 261 00 97', 'WordPress': 'Ja'},
    # Gynaeko Zürich
    {'Firma': 'Zürich Frauenarzt', 'Website': 'zuerichfrauenarzt.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '+41 44 311 48 66', 'WordPress': 'Ja'},
    {'Firma': 'Gynpoint', 'Website': 'gynpoint.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '+41 43 336 70 20', 'WordPress': 'Ja'},
    # Kardio Zürich
    {'Firma': 'DTH Herzzentrum Zürich', 'Website': 'dth-herzzentrum.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '', 'WordPress': 'Ja'},
    # Kardio Bern
    {'Firma': 'KGP Kardiologie Bern', 'Website': 'kgp-bern.ch', 'EMail': '', 'Stadt': 'Bern', 'Telefon': '', 'WordPress': 'Ja'},
]

def normalize_url(url):
    url = str(url).strip().lower().rstrip('/')
    url = url.replace('https://', '').replace('http://', '').replace('www.', '')
    return url

wb = openpyxl.load_workbook(input_file)
ws = wb['Praxen Leads']
existing_urls = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    val = str(row[0].value or '').strip()
    if val: existing_urls.add(normalize_url(val))
print(f"Bestehende Eintraege: {ws.max_row - 1}")

thin_border = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'),
                     top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))
data_font = Font(name='Calibri', size=11, color='333333')
data_alignment = Alignment(horizontal='left', vertical='center')
added = skipped = 0
for lead in new_leads:
    norm_url = normalize_url(lead['Website'])
    if norm_url in existing_urls:
        print(f"  SKIP: {lead['Website']}"); skipped += 1; continue
    next_row = ws.max_row + 1
    for i, key in enumerate(['Firma','Website','EMail','Stadt','Telefon','WordPress'], 1):
        cell = ws.cell(row=next_row, column=i, value=lead[key])
        cell.font = data_font; cell.alignment = data_alignment; cell.border = thin_border
    existing_urls.add(norm_url); added += 1
    print(f"  ADDED: {lead['Firma']} ({lead['Website']})")
wb.save(input_file)
print(f"\nErgebnis: {added} hinzugefuegt, {skipped} uebersprungen")
print(f"Gesamt 'Praxen Leads': {ws.max_row - 1} Eintraege")
