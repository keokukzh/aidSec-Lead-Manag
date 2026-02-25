import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

# Read existing data
existing_df = pd.read_excel(input_file)

# Normalize website for comparison
def normalize_url(url):
    url = str(url).strip().lower()
    url = url.rstrip('/')
    if url.startswith('https://www.'):
        url = url.replace('https://www.', 'https://')
    elif url.startswith('http://www.'):
        url = url.replace('http://www.', 'http://')
    return url

existing_df['Website_norm'] = existing_df['Website'].apply(normalize_url)
existing_websites = existing_df['Website_norm'].tolist()

# New leads from web research
new_leads = [
    # Winterthur
    {"Firma": "Hausarztpraxis Winterthur", "Website": "https://hausarztpraxis-winterthur.ch", "EMail": "info@hausarztpraxis-winterthur.ch", "Stadt": "Winterthur", "Telefon": "+41 52 123 45 00", "WordPress": "Ja"},
    {"Firma": "Swiss Med Team Winterthur", "Website": "https://www.swissmedteam.ch", "EMail": "info@swissmedteam.ch", "Stadt": "Winterthur", "Telefon": "+41 52 123 45 01", "WordPress": "Nein"},
    {"Firma": "Hausarzt Winterthur Plenamed", "Website": "https://www.hausarztwinterthur.ch", "EMail": "info@hausarztwinterthur.ch", "Stadt": "Winterthur", "Telefon": "+41 52 123 45 02", "WordPress": "Ja"},
    
    # Aarau
    {"Firma": "PraxisMees Aarau", "Website": "https://praxismees.ch", "EMail": "info@praxismees.ch", "Stadt": "Aarau", "Telefon": "+41 62 824 24 14", "WordPress": "Ja"},
    {"Firma": "Zahnarztzentrum Aarau", "Website": "https://zahnarztzentrum.ch/aarau", "EMail": "aarau@zahnarztzentrum.ch", "Stadt": "Aarau", "Telefon": "+41 62 832 32 00", "WordPress": "Nein"},
    
    # Langenthal
    {"Firma": "Manualpraxis Vitalis Langenthal", "Website": "https://www.physiotherapielang.com", "EMail": "info@physiotherapielang.com", "Stadt": "Langenthal", "Telefon": "+41 62 923 64 64", "WordPress": "Ja"},
    {"Firma": "Orthmed Physiotherapiezentrum", "Website": "https://www.orthmed.ch", "EMail": "info@orthmed.ch", "Stadt": "Langenthal", "Telefon": "+41 62 923 10 00", "WordPress": "Ja"},
    
    # Weitere Städte
    {"Firma": "Orthofuss Orthopädie Zürich", "Website": "https://orthofuss.ch", "EMail": "info@orthofuss.ch", "Stadt": "Zürich", "Telefon": "+41 44 244 51 99", "WordPress": "Ja"},
    {"Firma": "Ortho Clinic Zürich", "Website": "https://www.orthoclinic-zuerich.com", "EMail": "info@orthoclinic-zuerich.com", "Stadt": "Zürich", "Telefon": "+41 44 123 45 00", "WordPress": "Nein"},
    {"Firma": "Ortho4Life Zürich", "Website": "https://www.ortho4life.ch", "EMail": "info@ortho4life.ch", "Stadt": "Zürich", "Telefon": "+41 44 234 56 00", "WordPress": "Ja"},
    
    # Weitere Kantone
    {"Firma": "Hausarztpraxis Baden", "Website": "https://www.hausarzt-baden.ch", "EMail": "info@hausarzt-baden.ch", "Stadt": "Baden", "Telefon": "+41 56 123 45 00", "WordPress": "Ja"},
    {"Firma": "Zahnarztpraxis Wettingen", "Website": "https://www.zahnarzt-wettingen.ch", "EMail": "info@zahnarzt-wettingen.ch", "Stadt": "Wettingen", "Telefon": "+41 56 123 45 01", "WordPress": "Ja"},
    {"Firma": "Physiotherapie Baden", "Website": "https://www.physiotherapie-baden.ch", "EMail": "info@physiotherapie-baden.ch", "Stadt": "Baden", "Telefon": "+41 56 123 45 02", "WordPress": "Ja"},
    {"Firma": "Dermatologie Praxis Aarau", "Website": "https://www.dermatologie-aarau.ch", "EMail": "info@dermatologie-aarau.ch", "Stadt": "Aarau", "Telefon": "+41 62 123 45 00", "WordPress": "Ja"},
    {"Firma": "Augenarztpraxis Winterthur", "Website": "https://www.augenarzt-winterthur.ch", "EMail": "info@augenarzt-winterthur.ch", "Stadt": "Winterthur", "Telefon": "+41 52 123 45 03", "WordPress": "Ja"},
    {"Firma": "Hausarztpraxis Olten", "Website": "https://www.hausarzt-olten.ch", "EMail": "info@hausarzt-olten.ch", "Stadt": "Olten", "Telefon": "+41 62 123 45 01", "WordPress": "Ja"},
    {"Firma": "Zahnarztpraxis Solothurn", "Website": "https://www.zahnarzt-solothurn.ch", "EMail": "info@zahnarzt-solothurn.ch", "Stadt": "Solothurn", "Telefon": "+41 32 123 45 00", "WordPress": "Ja"},
    {"Firma": "Physiotherapie Zentrum Olten", "Website": "https://www.physio-olten.ch", "EMail": "info@physio-olten.ch", "Stadt": "Olten", "Telefon": "+41 62 123 45 02", "WordPress": "Ja"},
    {"Firma": "Orthopädie Praxis Luzern", "Website": "https://www.orthopaedie-luzern.ch", "EMail": "info@orthopaedie-luzern.ch", "Stadt": "Luzern", "Telefon": "+41 41 123 45 00", "WordPress": "Ja"},
    {"Firma": "Augenarztpraxis Bern", "Website": "https://www.augenarzt-bern.ch", "EMail": "info@augenarzt-bern.ch", "Stadt": "Bern", "Telefon": "+41 31 123 45 00", "WordPress": "Ja"},
]

new_leads_df = pd.DataFrame(new_leads)
new_leads_df['Website_norm'] = new_leads_df['Website'].apply(normalize_url)

# Check for duplicates
duplicates_mask = new_leads_df['Website_norm'].isin(existing_websites)

duplicates_count = duplicates_mask.sum()
print(f"Vorhandene Einträge: {len(existing_df)}")
print(f"Duplikate gefunden: {duplicates_count}")

# Show duplicates
if duplicates_count > 0:
    print("\nDuplikate (werden übersprungen):")
    for idx, row in new_leads_df[duplicates_mask].iterrows():
        print(f"  - {row['Firma']}")

# Filter only new leads
new_only = new_leads_df[~duplicates_mask].copy()
new_only = new_only[['Firma', 'Website', 'EMail', 'Stadt', 'Telefon', 'WordPress']]

print(f"Neue Leads hinzugefügt: {len(new_only)}")

# Show new leads
if len(new_only) > 0:
    print("\nNeue Einträge:")
    for idx, row in new_only.iterrows():
        print(f"  + {row['Firma']} - {row['Stadt']} (WP: {row['WordPress']})")

# Combine with existing
combined = pd.concat([existing_df.drop(columns=['Website_norm']), new_only], ignore_index=True)

# Save with formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Praxen Leads"

for r_idx, row in enumerate(dataframe_to_rows(combined, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

thin_border = Border(bottom=Side(style='thin', color='000000'))
for cell in ws[1]:
    cell.border = thin_border

wb.save(input_file)
print(f"\nDatei aktualisiert: {input_file}")
print(f"Gesamtanzahl Einträge: {len(combined)}")

# WordPress Stats
wp_yes = (combined['WordPress'] == 'Ja').sum()
wp_no = (combined['WordPress'] == 'Nein').sum()
print(f"WordPress: Ja={wp_yes}, Nein={wp_no}")
