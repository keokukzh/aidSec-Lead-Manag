import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

new_leads = [
    # Chur
    {'Firma': 'CHURMED Praxisgemeinschaft', 'Website': 'churmed.ch', 'EMail': '', 'Stadt': 'Chur', 'Telefon': '+41 81 253 70 70', 'WordPress': 'Ja'},
    # Schaffhausen
    {'Firma': 'Dentalzentrum Schaffhausen', 'Website': 'zahnarzt-dental-schaffhausen.ch', 'EMail': '', 'Stadt': 'Schaffhausen', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Zahnarzt Dr. Antonescu', 'Website': 'schaffhausenzahnarzt.ch', 'EMail': '', 'Stadt': 'Schaffhausen', 'Telefon': '', 'WordPress': 'Ja'},
    # Thun
    {'Firma': 'Hausarztpraxis Thun', 'Website': 'hausarztpraxis-thun.ch', 'EMail': '', 'Stadt': 'Thun', 'Telefon': '+41 33 244 11 44', 'WordPress': 'Ja'},
    # Buelach
    {'Firma': 'Lifedent Bülach', 'Website': 'lifedent.ch', 'EMail': '', 'Stadt': 'Bülach', 'Telefon': '+41 44 860 45 45', 'WordPress': 'Ja'},
    # Rapperswil
    {'Firma': 'rappjderm AG', 'Website': 'rappjderm.ch', 'EMail': '', 'Stadt': 'Rapperswil-Jona', 'Telefon': '', 'WordPress': 'Ja'},
    # Baselland
    {'Firma': 'Smile Clinix', 'Website': 'smileclinix.ch', 'EMail': '', 'Stadt': 'Liestal', 'Telefon': '+41 61 927 13 13', 'WordPress': 'Ja'},
    {'Firma': 'Zahnarztpraxis Müller & Lorant', 'Website': 'zahnarzt-muttenz.ch', 'EMail': '', 'Stadt': 'Muttenz', 'Telefon': '+41 61 461 52 51', 'WordPress': 'Ja'},
    {'Firma': 'Zahnklinik Rennbahn', 'Website': 'zahnklinik-rennbahn.ch', 'EMail': '', 'Stadt': 'Muttenz', 'Telefon': '+41 61 826 10 10', 'WordPress': 'Ja'},
    # Schlieren/Limmattal
    {'Firma': 'Doktorhuus Schlieren', 'Website': 'schlieren.doktor-huus.ch', 'EMail': '', 'Stadt': 'Schlieren', 'Telefon': '+41 44 755 44 44', 'WordPress': 'Ja'},
    # Wetzikon/Zuercher Oberland
    {'Firma': 'Zahnteam Wetzikon', 'Website': 'zahnteam-wetzikon.ch', 'EMail': '', 'Stadt': 'Wetzikon', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Zahnarztpraxis Volketswil', 'Website': 'zahnarzt-volketswil.ch', 'EMail': '', 'Stadt': 'Volketswil', 'Telefon': '+41 44 945 57 81', 'WordPress': 'Ja'},
    {'Firma': 'Zahnarztpraxis Easydent', 'Website': 'zahnarztpraxis-easydent.ch', 'EMail': '', 'Stadt': 'Wetzikon', 'Telefon': '', 'WordPress': 'Ja'},
    # Liestal
    {'Firma': 'Hausarztpraxis Rheinstrasse', 'Website': 'hausarztpraxis-rheinstrasse.ch', 'EMail': 'praxis-rheinstrasse@hin.ch', 'Stadt': 'Liestal', 'Telefon': '+41 61 517 50 70', 'WordPress': 'Ja'},
    # Luzern Physio
    {'Firma': 'Convalis Physiotherapie', 'Website': 'convalis-physiotherapie.ch', 'EMail': '', 'Stadt': 'Kriens', 'Telefon': '+41 41 310 32 70', 'WordPress': 'Ja'},
    {'Firma': "Piet's Praxis", 'Website': 'pietspraxis.ch', 'EMail': '', 'Stadt': 'Luzern', 'Telefon': '+41 78 227 50 91', 'WordPress': 'Ja'},
    # Chur Zahnarzt
    {'Firma': 'SmilePraxis', 'Website': 'smilepraxis.ch', 'EMail': '', 'Stadt': 'Chur', 'Telefon': '+41 81 322 40 00', 'WordPress': 'Ja'},
]

def normalize_url(url):
    url = str(url).strip().lower().rstrip('/')
    url = url.replace('https://', '').replace('http://', '').replace('www.', '')
    return url

wb = openpyxl.load_workbook(input_file)
ws = wb['Praxen Leads']
existing_urls = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    val = str(row[0].value or '').strip()
    if val: existing_urls.add(normalize_url(val))
print(f"Bestehende Eintraege: {ws.max_row - 1}")

thin_border = Border(left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'),
                     top=Side(style='thin', color='E0E0E0'), bottom=Side(style='thin', color='E0E0E0'))
data_font = Font(name='Calibri', size=11, color='333333')
data_alignment = Alignment(horizontal='left', vertical='center')
added = skipped = 0
for lead in new_leads:
    norm_url = normalize_url(lead['Website'])
    if norm_url in existing_urls:
        print(f"  SKIP: {lead['Website']}"); skipped += 1; continue
    next_row = ws.max_row + 1
    for i, key in enumerate(['Firma','Website','EMail','Stadt','Telefon','WordPress'], 1):
        cell = ws.cell(row=next_row, column=i, value=lead[key])
        cell.font = data_font; cell.alignment = data_alignment; cell.border = thin_border
    existing_urls.add(norm_url); added += 1
    print(f"  ADDED: {lead['Firma']} ({lead['Website']})")
wb.save(input_file)
print(f"\nErgebnis: {added} hinzugefuegt, {skipped} uebersprungen")
print(f"Gesamt 'Praxen Leads': {ws.max_row - 1} Eintraege")
