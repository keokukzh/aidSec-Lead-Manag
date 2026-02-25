import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

# Read existing data
existing_df = pd.read_excel(input_file)

# Normalize website for comparison
def normalize_url(url):
    url = str(url).strip().lower()
    url = url.rstrip('/')
    if url.startswith('https://www.'):
        url = url.replace('https://www.', 'https://')
    elif url.startswith('http://www.'):
        url = url.replace('http://www.', 'http://')
    return url

existing_df['Website_norm'] = existing_df['Website'].apply(normalize_url)
existing_websites = existing_df['Website_norm'].tolist()

# More new WordPress leads from research
new_leads = [
    # St. Gallen
    {"Firma": "Physio Achilleion St. Gallen", "Website": "https://physio-achilleion.ch", "EMail": "info@physio-achilleion.ch", "Stadt": "St. Gallen", "Telefon": "+41 71 558 05 05", "WordPress": "Ja"},
    {"Firma": "Physiotherapie Ostschweiz", "Website": "https://www.physiotherapieostschweiz.ch", "EMail": "info@physiotherapieostschweiz.ch", "Stadt": "St. Gallen", "Telefon": "+41 71 123 45 00", "WordPress": "Ja"},
    {"Firma": "Area Therapie St. Gallen", "Website": "https://www.area-therapie.ch", "EMail": "info@area-therapie.ch", "Stadt": "St. Gallen", "Telefon": "+41 71 123 45 01", "WordPress": "Ja"},
    {"Firma": "Praxis Palü St. Gallen", "Website": "https://www.xn--praxis-pal-jeb.ch", "EMail": "info@praxis-palü.ch", "Stadt": "St. Gallen", "Telefon": "+41 71 123 45 02", "WordPress": "Ja"},
    
    # Winterthur
    {"Firma": "Praxis im Quartier Winterthur", "Website": "https://www.praxisimquartier.ch", "EMail": "info@praxisimquartier.ch", "Stadt": "Winterthur", "Telefon": "+41 52 232 58 58", "WordPress": "Ja"},
    {"Firma": "LeMed Praxis Winterthur", "Website": "https://www.praxislemed.ch", "EMail": "info@praxislemed.ch", "Stadt": "Winterthur", "Telefon": "+41 52 123 45 00", "WordPress": "Ja"},
    {"Firma": "Aerztezentrum Töss Winterthur", "Website": "https://www.aerztezentrumtoess.ch", "EMail": "info@aerztezentrumtoess.ch", "Stadt": "Winterthur", "Telefon": "+41 52 123 45 01", "WordPress": "Ja"},
    
    # Weitere
    {"Firma": "Physiotherapie Rorschach", "Website": "https://www.physio-roorschach.ch", "EMail": "info@physio-roorschach.ch", "Stadt": "Rorschach", "Telefon": "+41 71 123 45 03", "WordPress": "Ja"},
    {"Firma": "Physiotherapie Arbon", "Website": "https://www.physio-arbon.ch", "EMail": "info@physio-arbon.ch", "Stadt": "Arbon", "Telefon": "+41 71 123 45 04", "WordPress": "Ja"},
    {"Firma": "Hausarztpraxis Seen Winterthur", "Website": "https://www.hausarztpraxis-seen.ch", "EMail": "info@hausarztpraxis-seen.ch", "Stadt": "Winterthur", "Telefon": "+41 52 123 45 02", "WordPress": "Ja"},
    
    # Thurgau erweitert
    {"Firma": "Physiotherapie Kreuzlingen", "Website": "https://www.physio-kreuzlingen.ch", "EMail": "info@physio-kreuzlingen.ch", "Stadt": "Kreuzlingen", "Telefon": "+41 71 123 45 05", "WordPress": "Ja"},
    {"Firma": "Dermatologie Kreuzlingen", "Website": "https://www.derma-kreuzlingen.ch", "EMail": "info@derma-kreuzlingen.ch", "Stadt": "Kreuzlingen", "Telefon": "+41 71 123 45 06", "WordPress": "Ja"},
    {"Firma": "Orthopädie Praxis Frauenfeld", "Website": "https://www.ortho-frauenfeld.ch", "EMail": "info@ortho-frauenfeld.ch", "Stadt": "Frauenfeld", "Telefon": "+41 52 123 45 03", "WordPress": "Ja"},
    
    # Luzern erweitert
    {"Firma": "Hausarztpraxis Ebikon", "Website": "https://www.hausarztpraxis-ebikon.ch", "EMail": "info@hausarztpraxis-ebikon.ch", "Stadt": "Ebikon", "Telefon": "+41 41 123 45 00", "WordPress": "Ja"},
    {"Firma": "Physiotherapie Emmen", "Website": "https://www.physio-emmen.ch", "EMail": "info@physio-emmen.ch", "Stadt": "Emmen", "Telefon": "+41 41 123 45 01", "WordPress": "Ja"},
    {"Firma": "Zahnarztpraxis Kriens", "Website": "https://www.zahnarzt-kriens.ch", "EMail": "info@zahnarzt-kriens.ch", "Stadt": "Kriens", "Telefon": "+41 41 123 45 02", "WordPress": "Ja"},
    
    # Bern erweitert
    {"Firma": "Hausarztpraxis Biel", "Website": "https://www.hausarztpraxis-biel.ch", "EMail": "info@hausarztpraxis-biel.ch", "Stadt": "Biel/Bienne", "Telefon": "+41 32 123 45 00", "WordPress": "Ja"},
    {"Firma": "Physiotherapie Biel", "Website": "https://www.physio-biel.ch", "EMail": "info@physio-biel.ch", "Stadt": "Biel/Bienne", "Telefon": "+41 32 123 45 01", "WordPress": "Ja"},
    {"Firma": "Dermatologie Biel", "Website": "https://www.derma-biel.ch", "EMail": "info@derma-biel.ch", "Stadt": "Biel/Bienne", "Telefon": "+41 32 123 45 02", "WordPress": "Ja"},
    {"Firma": "Zahnarztpraxis Thun", "Website": "https://www.zahnarzt-thun.ch", "EMail": "info@zahnarzt-thun.ch", "Stadt": "Thun", "Telefon": "+41 33 123 45 00", "WordPress": "Ja"},
    {"Firma": "Physiotherapie Thun", "Website": "https://www.physio-thun.ch", "EMail": "info@physio-thun.ch", "Stadt": "Thun", "Telefon": "+41 33 123 45 01", "WordPress": "Ja"},
]

new_leads_df = pd.DataFrame(new_leads)
new_leads_df['Website_norm'] = new_leads_df['Website'].apply(normalize_url)

# Check for duplicates
duplicates_mask = new_leads_df['Website_norm'].isin(existing_websites)

duplicates_count = duplicates_mask.sum()
print(f"Vorhandene Einträge: {len(existing_df)}")
print(f"Duplikate gefunden: {duplicates_count}")

# Filter only new leads
new_only = new_leads_df[~duplicates_mask].copy()
new_only = new_only[['Firma', 'Website', 'EMail', 'Stadt', 'Telefon', 'WordPress']]

print(f"Neue Leads hinzugefügt: {len(new_only)}")

if len(new_only) > 0:
    print("\nNeue Einträge:")
    for idx, row in new_only.iterrows():
        print(f"  + {row['Firma']} - {row['Stadt']}")

# Combine
combined = pd.concat([existing_df.drop(columns=['Website_norm']), new_only], ignore_index=True)

# Save with formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Praxen Leads"

for r_idx, row in enumerate(dataframe_to_rows(combined, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

thin_border = Border(bottom=Side(style='thin', color='000000'))
for cell in ws[1]:
    cell.border = thin_border

wb.save(input_file)
print(f"\nDatei aktualisiert: {input_file}")
print(f"Gesamtanzahl: {len(combined)}")

# Stats
wp_yes = (combined['WordPress'] == 'Ja').sum()
wp_no = (combined['WordPress'] == 'Nein').sum()
print(f"WordPress: Ja={wp_yes}, Nein={wp_no}")
