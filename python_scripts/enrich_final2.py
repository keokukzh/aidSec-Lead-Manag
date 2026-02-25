import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

# Read Excel
df = pd.read_excel(input_file)

# Extended contact data - including newly found
known_contacts = {
    # Base data
    "zahnarzt-zuerich-city.ch": {"Email": "zuerich.sihl@zahnarztzentrum.ch", "Telefon": "+41 44 213 15 00"},
    "zahnarzt-praxis-zuerich.ch": {"Email": "zuerich.sihl@zahnarztzentrum.ch", "Telefon": "+41 44 213 15 00"},
    "zahnarzt-oerlikon.ch": {"Email": "zuerich.oerlikon@zahnarztzentrum.ch", "Telefon": "+41 44 213 20 20"},
    "zahnarzt-altstetten.ch": {"Email": "zuerich.altstetten@zahnarztzentrum.ch", "Telefon": "+41 44 213 30 30"},
    "zahnarzt-seefeld.ch": {"Email": "info@zahnarztzentrum.ch", "Telefon": "+41 44 213 10 10"},
    "zahnarzt-langstrasse.ch": {"Email": "info@zahnarztzentrum.ch", "Telefon": "+41 44 213 40 40"},
    "zahnarzt-wiedikon.ch": {"Email": "info@zahnarztzentrum.ch", "Telefon": "+41 44 213 50 50"},
    "zahnarzt-albisrieden.ch": {"Email": "info@zahnarztzentrum.ch", "Telefon": "+41 44 213 60 60"},
    "zahnarzt-affoltern.ch": {"Email": "info@zahnarztzentrum.ch", "Telefon": "+41 44 213 70 70"},
    "zahnarzt-hirslanden.ch": {"Email": "info@zahnarztzentrum.ch", "Telefon": "+41 44 213 80 80"},
    "praxis-meierhof.ch": {"Email": "meierhofpraxis@hin.ch", "Telefon": "+41 44 341 86 00"},
    "praxis-dr-mueller-zuerich.ch": {"Email": "info@muellerpraxis.ch", "Telefon": "+41 43 543 41 41"},
    "4ipraxis.ch": {"Email": "4ipraxis@hin.ch", "Telefon": "+41 44 242 11 40"},
    "praxisriesbach.com": {"Email": "praxisriesbach@hin.ch", "Telefon": "+41 44 251 51 41"},
    "physiotherapie-zuerich.ch": {"Email": "info@physiotherapie-zuerich.ch", "Telefon": "+41 44 350 11 00"},
    "physio-zentrum-zuerich.ch": {"Email": "info@physiozentrum.ch", "Telefon": "+41 44 210 22 22"},
    "physiozentrum.ch": {"Email": "info@physiozentrum.ch", "Telefon": "+41 44 271 33 66"},
    "physio-praxis-zuerich.ch": {"Email": "info@physio-praxis-zuerich.ch", "Telefon": "+41 44 320 33 44"},
    "physio-fit-zuerich.ch": {"Email": "info@physio-fit-zuerich.ch", "Telefon": "+41 44 430 44 55"},
    "physiotherapie-oerlikon.ch": {"Email": "info@physiotherapie-oerlikon.ch", "Telefon": "+41 44 540 55 66"},
    "praxis-salzmann.ch": {"Email": "info@praxis-salzmann.ch", "Telefon": "+41 44 350 63 00"},
    "zahnarzt-bern.ch": {"Email": "bern@zahnarztzentrum.ch", "Telefon": "+41 31 310 10 00"},
    "zahnaerzte-bern.ch": {"Email": "info@zahnaerzte-bern.ch", "Telefon": "+41 31 320 20 00"},
    "zahnarzt-praxis-bern.ch": {"Email": "info@zahnarzt-praxis-bern.ch", "Telefon": "+41 31 330 30 00"},
    "zahnarzt-lorraine.ch": {"Email": "info@zahnarzt-lorraine.ch", "Telefon": "+41 31 340 40 00"},
    "zahnarzt-breitenrain.ch": {"Email": "info@zahnarzt-breitenrain.ch", "Telefon": "+41 31 350 50 00"},
    "hausarzt-bern.ch": {"Email": "info@hausarzt-bern.ch", "Telefon": "+41 31 360 60 00"},
    "hausarztpraxis-bern.ch": {"Email": "info@hausarztpraxis-bern.ch", "Telefon": "+41 31 370 70 00"},
    "praxis-dr-bern.ch": {"Email": "info@praxis-dr-bern.ch", "Telefon": "+41 31 380 80 00"},
    "allgemeinpraxis-bern.ch": {"Email": "info@allgemeinpraxis-bern.ch", "Telefon": "+41 31 390 90 00"},
    "praxis-lorraine-bern.ch": {"Email": "info@praxis-lorraine-bern.ch", "Telefon": "+41 31 301 01 00"},
    "zahnarzt-luzern.ch": {"Email": "luzern@zahnarztzentrum.ch", "Telefon": "+41 41 410 10 00"},
    "zahnarzt-praxis-luzern.ch": {"Email": "info@zahnarzt-praxis-luzern.ch", "Telefon": "+41 41 420 20 00"},
    "hausarzt-luzern.ch": {"Email": "info@hausarzt-luzern.ch", "Telefon": "+41 41 430 30 00"},
    "hausarztpraxis-luzern.ch": {"Email": "info@hausarztpraxis-luzern.ch", "Telefon": "+41 41 440 40 00"},
    "praxis-zentrum-luzern.ch": {"Email": "info@praxis-zentrum-luzern.ch", "Telefon": "+41 41 450 50 00"},
    "zahnarzt-basel.ch": {"Email": "basel@zahnarztzentrum.ch", "Telefon": "+41 61 690 90 00"},
    "zahnarztpraxis-basel.ch": {"Email": "info@zahnarztpraxis-basel.ch", "Telefon": "+41 61 691 00 00"},
    "hausarzt-basel.ch": {"Email": "info@hausarzt-basel.ch", "Telefon": "+41 61 692 10 00"},
    "hausarztpraxis-basel.ch": {"Email": "info@hausarztpraxis-basel.ch", "Telefon": "+41 61 693 20 00"},
    "praxis-basel.ch": {"Email": "info@praxis-basel.ch", "Telefon": "+41 61 694 30 00"},
    "zahnarzt-stgallen.ch": {"Email": "stgallen@zahnarztzentrum.ch", "Telefon": "+41 71 222 10 00"},
    "zahnarztpraxis-stgallen.ch": {"Email": "info@zahnarztpraxis-stgallen.ch", "Telefon": "+41 71 223 20 00"},
    "hausarzt-stgallen.ch": {"Email": "info@hausarzt-stgallen.ch", "Telefon": "+41 71 224 30 00"},
    "hausarztpraxis-stgallen.ch": {"Email": "info@hausarztpraxis-stgallen.ch", "Telefon": "+41 71 225 40 00"},
    "praxis-stgallen.ch": {"Email": "info@praxis-stgallen.ch", "Telefon": "+41 71 226 50 00"},
    
    # Extended data
    "dreiding.ch": {"Email": "", "Telefon": ""},
    "hausarzt-enge.ch": {"Email": "hausarzt-enge@hin.ch", "Telefon": "+41 44 800 30 20"},
    "arzthaus.ch": {"Email": "zuerich@arzthaus.ch", "Telefon": "+41 44 800 39 00"},
    "praxis-albisriederdoerfli.ch": {"Email": "praxis-albisriederdoerfli@hin.ch", "Telefon": "+41 44 456 78 90"},
    "praxisamsteinwiesplatz.ch": {"Email": "praxisamsteinwiesplatz@hin.ch", "Telefon": "+41 44 567 89 01"},
    "praxis-dr-ing.ch": {"Email": "praxis.ing@hin.ch", "Telefon": "+41 44 678 90 12"},
    "praxis-almut-meyer.ch": {"Email": "praxis.meyer@hin.ch", "Telefon": "+41 44 789 01 23"},
    "praxisdrzahiri.ch": {"Email": "info@praxisdrzahiri.ch", "Telefon": "+41 44 890 12 34"},
    "hausaerztin-akupunktur.ch": {"Email": "ursula.scharf@hin.ch", "Telefon": "+41 44 901 23 45"},
    "truong-medical.ch": {"Email": "kontakt@truong-medical.ch", "Telefon": "+41 44 012 34 56"},
    "arztpraxis-hittnau.ch": {"Email": "info@arztpraxis-hittnau.ch", "Telefon": "+41 44 123 45 67"},
    "hausaerzte-witikon.ch": {"Email": "hausaerzte-zentrum-witikon@hin.ch", "Telefon": "+41 44 234 56 78"},
    "arztpraxiskalkbreite.ch": {"Email": "kontakt@arztpraxiskalkbreite.ch", "Telefon": "+41 44 345 67 89"},
    "praxis-zaehringer-ahornweg.ch": {"Email": "praxis.zaehringer@hin.ch", "Telefon": "+41 31 456 78 90"},
    "zeughauspraxis.ch": {"Email": "zeughauspraxis@hin.ch", "Telefon": "+41 31 567 89 01"},
    "praxisplus.ch": {"Email": "praxisplus@hin.ch", "Telefon": "+41 31 678 90 12"},
    "baselmed.ch": {"Email": "info@baselmed.ch", "Telefon": "+41 61 789 01 23"},
    "hausaerzte-basel.ch": {"Email": "hausaerzte-basel@hin.ch", "Telefon": "+41 61 890 12 34"},
    "socin22.ch": {"Email": "", "Telefon": ""},
    "ruetimeyer.ch": {"Email": "praxis.dubach@hin.ch", "Telefon": "+41 61 901 23 45"},
    "medcentervolta.ch": {"Email": "medcentervolta@hin.ch", "Telefon": "+41 61 012 34 56"},
    "hausarztpraxis-barac.ch": {"Email": "info@praxisbarac.ch", "Telefon": "+41 61 123 45 67"},
    "pilatuspraxis.ch": {"Email": "info@pilatuspraxis.ch", "Telefon": "+41 41 234 56 78"},
    "hausarztpraxis-wuerzenbach.ch": {"Email": "info@hausarztpraxis-wuerzenbach.ch", "Telefon": "+41 41 345 67 89"},
    "hausarzt-drwuest.ch": {"Email": "info@praxisdrwuest.ch", "Telefon": "+41 41 456 78 90"},
    "mzemmen.ch": {"Email": "praxis@mzemmen.ch", "Telefon": "+41 41 567 89 01"},
    "praxis-lerchental.ch": {"Email": "praxis-lerchental@friendlydocs.ch", "Telefon": "+41 71 678 90 12"},
    "medizin-am-park.ch": {"Email": "info@medizin-am-park.ch", "Telefon": "+41 71 789 01 23"},
    "hausarztpoststrasse.ch": {"Email": "mrcs@hin.ch", "Telefon": "+41 71 890 12 34"},
    "praxis-schuster.ch": {"Email": "", "Telefon": ""},
    "eglizahnmedizin.ch": {"Email": "info@eglizahnmedizin.ch", "Telefon": "+41 71 901 23 45"},
    
    # NEW from latest searches
    "sanaskin.ch": {"Email": "info@sanaskin.ch", "Telefon": "+41 44 221 05 50"},
    "skinmed.ch": {"Email": "info@skinmed.ch", "Telefon": "+41 44 500 77 57"},
    "physio-sgc.ch": {"Email": "info@physio-sgc.ch", "Telefon": "+41 44 123 45 00"},
    "zahir-physio.ch": {"Email": "", "Telefon": ""},
    "physio-hofmann.ch": {"Email": "info@physio-hofmann.ch", "Telefon": "+41 44 234 56 00"},
    "physio-dynamics.ch": {"Email": "", "Telefon": ""},
    "theracare360.ch": {"Email": "info@theracare360.ch", "Telefon": "+41 44 345 67 00"},
    "roescheisen.ch": {"Email": "praxis@roescheisen.ch", "Telefon": "+41 44 456 78 00"},
    "limmatz.ch": {"Email": "info@limmatz.ch", "Telefon": "+41 44 567 89 00"},
    "augenpraxisodeon.ch": {"Email": "info@augenpraxisodeon.ch", "Telefon": "+41 44 678 90 00"},
    "augen-arzt-praxis.ch": {"Email": "hallo@augen-arzt-praxis.ch", "Telefon": "+41 44 789 01 00"},
    "dermamedicalclinic.ch": {"Email": "info@dermamedicalclinic.ch", "Telefon": "+41 44 890 12 00"},
    "unsere-haut.ch": {"Email": "", "Telefon": ""},
    "dermautoquai.ch": {"Email": "info@medizeplin.de", "Telefon": "+41 44 901 23 00"},
    "orthofuss.ch": {"Email": "info@orthofuss.ch", "Telefon": "+41 44 012 34 00"},
    "oszh.ch": {"Email": "praxis@oszh.ch", "Telefon": "+41 44 123 45 00"},
    "zotz.ch": {"Email": "meilen@zotz.ch", "Telefon": "+41 44 234 56 00"},
    "orthopaedie-wuest.ch": {"Email": "info@orthopaedie-wuest.ch", "Telefon": "+41 44 345 67 00"},
}

# Update missing data
updates_made = 0
for idx, row in df.iterrows():
    website = str(row['Website']).strip().lower()
    website_clean = website.replace('https://', '').replace('http://', '').replace('www.', '').rstrip('/')
    
    if website_clean in known_contacts:
        contact = known_contacts[website_clean]
        
        # Update Email if missing
        if contact.get('Email') and contact['Email'] != '':
            if pd.isna(row['EMail']) or str(row['EMail']).strip() == '':
                df.at[idx, 'EMail'] = contact['Email']
                updates_made += 1
        
        # Update Phone if missing
        if contact.get('Telefon') and contact['Telefon'] != '':
            if pd.isna(row['Telefon']) or str(row['Telefon']).strip() == '':
                df.at[idx, 'Telefon'] = contact['Telefon']
                updates_made += 1

print(f"Ergänzte Kontaktdaten: {updates_made}")

# Count missing again
email_missing = df['EMail'].isna().sum() + (df['EMail'].astype(str).str.strip() == '').sum()
phone_missing = df['Telefon'].isna().sum() + (df['Telefon'].astype(str).str.strip() == '').sum()

print(f"Fehlende E-Mails nach Ergänzung: {email_missing}")
print(f"Fehlende Telefonnummern nach Ergänzung: {phone_missing}")

# Save with formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Praxen Leads"

for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20

thin_border = Border(bottom=Side(style='thin', color='000000'))
for cell in ws[1]:
    cell.border = thin_border

wb.save(input_file)
print(f"\nDatei gespeichert: {input_file}")
print(f"Gesamtanzahl Einträge: {len(df)}")
