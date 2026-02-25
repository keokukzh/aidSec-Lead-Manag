import openpyxl

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"
wb = openpyxl.load_workbook(input_file)
ws = wb["Anwalts Kanzleien"]

print("=== FEHLENDE E-MAILS ===")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
    email = row[2].value
    if not email or str(email).strip() == '':
        print(f"  {row[0].value} | {row[1].value}")

print("\n=== FEHLENDE TELEFONE ===")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
    tel = row[4].value
    if not tel or str(tel).strip() == '':
        print(f"  {row[0].value} | {row[1].value}")

print("\n=== FEHLENDE STÄDTE ===")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
    stadt = row[3].value
    if not stadt or str(stadt).strip() == '':
        print(f"  {row[0].value} | {row[1].value}")
