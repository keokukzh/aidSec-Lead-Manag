import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re

# CSV file
csv_file = r"c:\Users\aidevelo\Desktop\Praxen leads\Anwaltskanzleien Deutschschweiz - WordPress Security Leads - Tabellenblatt1 (2).csv"

# Read CSV
df = pd.read_csv(csv_file, header=0)
print(f"CSV gelesen: {len(df)} Zeilen")
print(f"Spalten: {list(df.columns)}")

# Clean and filter data
# Rename columns to our format
df = df.rename(columns={
    'Name des Unternehmens': 'Firma',
    'Website-URL': 'Website',
    'Email': 'EMail',
    'Stadt': 'Stadt',
    'WordPress erkannt?': 'WordPress'
})

# Add Telefon column (empty for now)
df['Telefon'] = ''

# Keep only relevant columns
df = df[['Firma', 'Website', 'EMail', 'Stadt', 'Telefon', 'WordPress']]

# Clean Website - fix malformed URLs
def clean_website(url):
    if pd.isna(url):
        return ''
    url = str(url).strip()
    # Fix URLs starting with comma or dots
    url = re.sub(r'^[\.\,]+', '', url)
    url = url.strip()
    if url and not url.startswith('http'):
        url = 'https://' + url
    return url

df['Website'] = df['Website'].apply(clean_website)

# Filter: only WordPress = JA or entries with valid website
df_filtered = df[
    (df['WordPress'].astype(str).str.upper().str.contains('JA', na=False)) |
    (df['Website'].str.contains('wordpress', case=False, na=False)) |
    (df['Website'].str.startswith('https://', na=False))
].copy()

# Also include entries that have a valid website even if WordPress is empty
df_valid = df[df['Website'].str.startswith('https://', na=False)].copy()

df_final = df_valid.copy()

# Remove duplicates based on website
df_final = df_final.drop_duplicates(subset=['Website'], keep='first')

# Remove empty websites
df_final = df_final[df_final['Website'].str.strip() != '']

# Set WordPress to "Ja" for all (since this is the WP leads list)
df_final['WordPress'] = 'Ja'

print(f"\nNach Bereinigung: {len(df_final)} Einträge")

# Clean EMail
def clean_email(email):
    if pd.isna(email):
        return ''
    email = str(email).strip()
    email = email.strip('()')
    return email

df_final['EMail'] = df_final['EMail'].apply(clean_email)

# Clean Stadt
def clean_city(city):
    if pd.isna(city):
        return ''
    city = str(city).strip()
    return city

df_final['Stadt'] = df_final['Stadt'].apply(clean_city)

# Show first entries
print("\nErste 10 Einträge:")
for idx, row in df_final.head(10).iterrows():
    print(f"  - {row['Firma'][:40]:<40} | {row['Website'][:30]:<30}")

# Save to Excel with both sheets
output_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

# Load existing workbook
wb = openpyxl.load_workbook(output_file)

# Create new sheet
ws = wb.create_sheet("Anwalts Kanzleien")

# Write header and data
for r_idx, row in enumerate(dataframe_to_rows(df_final.reset_index(drop=True), index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

# Column widths
ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 18
ws.column_dimensions['F'].width = 12

# Header border
thin_border = Border(bottom=Side(style='thin', color='000000'))
for cell in ws[1]:
    cell.border = thin_border

# Save
wb.save(output_file)
print(f"\nDatei gespeichert: {output_file}")
print(f"Anwalts Kanzleien: {len(df_final)} Einträge")
