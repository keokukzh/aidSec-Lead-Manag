import openpyxl

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"
wb = openpyxl.load_workbook(input_file)
ws = wb["Anwalts Kanzleien"]

# Additional data found in second research pass
additional_data = {
    'steinlex.ch': {
        'Telefon': '+41 44 269 40 00',
    },
    'advokatur-zuerich.ch': {
        'Firma': 'Fischer Rechtsanwälte',
        'Telefon': '+41 44 515 56 56',
        'EMail': 'sekretariat@fischer-rechtsanwaelte.ch',
    },
}

updates = 0
for row_idx in range(2, ws.max_row + 1):
    website = str(ws.cell(row=row_idx, column=2).value or '').lower()

    for domain, data in additional_data.items():
        if domain in website:
            for field, value in data.items():
                col_map = {'Firma': 1, 'EMail': 3, 'Stadt': 4, 'Telefon': 5}
                col = col_map.get(field)
                if col:
                    current = ws.cell(row=row_idx, column=col).value
                    if not current or str(current).strip() == '':
                        ws.cell(row=row_idx, column=col).value = value
                        updates += 1
                        print(f"  Update: {ws.cell(row=row_idx, column=1).value} -> {field} = {value}")

print(f"\n{updates} Felder aktualisiert")

# Final statistics
print(f"\n=== FINALE STATISTIKEN ===")
missing_email = 0
missing_phone = 0
missing_city = 0
total = 0

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
    total += 1
    if not row[2].value or str(row[2].value).strip() == '':
        missing_email += 1
    if not row[4].value or str(row[4].value).strip() == '':
        missing_phone += 1
    if not row[3].value or str(row[3].value).strip() == '':
        missing_city += 1

print(f"Gesamt: {total} Einträge")
print(f"E-Mail:   {total - missing_email}/{total} ({(total-missing_email)/total*100:.0f}%)")
print(f"Telefon:  {total - missing_phone}/{total} ({(total-missing_phone)/total*100:.0f}%)")
print(f"Stadt:    {total - missing_city}/{total} ({(total-missing_city)/total*100:.0f}%)")

if missing_email > 0:
    print(f"\nFehlende E-Mails ({missing_email}):")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        if not row[2].value or str(row[2].value).strip() == '':
            print(f"  - {row[0].value} ({row[1].value})")

if missing_phone > 0:
    print(f"\nFehlende Telefone ({missing_phone}):")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        if not row[4].value or str(row[4].value).strip() == '':
            print(f"  - {row[0].value} ({row[1].value})")

if missing_city > 0:
    print(f"\nFehlende Städte ({missing_city}):")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        if not row[3].value or str(row[3].value).strip() == '':
            print(f"  - {row[0].value} ({row[1].value})")

wb.save(input_file)
print(f"\nDatei gespeichert: {input_file}")
