import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

new_leads = [
    {'Firma': 'AdvoBern', 'Website': 'advo-bern.ch', 'EMail': 'info@advo-bern.ch', 'Stadt': 'Bern', 'Telefon': '+41 31 311 75 36', 'WordPress': 'Ja'},
    {'Firma': 'Advokatur & Notariat Bern', 'Website': 'anb-recht.ch', 'EMail': 'kontakt@anb-recht.ch', 'Stadt': 'Bern', 'Telefon': '+41 31 382 02 72', 'WordPress': 'Ja'},
    {'Firma': 'ADVOTECH ADVOKATEN', 'Website': 'advotech.ch', 'EMail': 'advotech@advotech.ch', 'Stadt': 'Basel', 'Telefon': '+41 61 205 48 00', 'WordPress': 'Ja'},
    {'Firma': 'WALDMANN Rechtsanwälte & Notare', 'Website': 'lawyers.ch', 'EMail': 'info@lawyers.ch', 'Stadt': 'Basel', 'Telefon': '+41 61 269 50 60', 'WordPress': 'Ja'},
    {'Firma': 'Advoschmid', 'Website': 'advoschmid.ch', 'EMail': 'mail@advoschmid.ch', 'Stadt': 'Solothurn', 'Telefon': '+41 32 622 58 21', 'WordPress': 'Ja'},
    {'Firma': 'Kanzlei Pilatushof', 'Website': 'pilatushof.ch', 'EMail': 'kanzlei@pilatushof.ch', 'Stadt': 'Luzern', 'Telefon': '+41 41 226 61 00', 'WordPress': 'Ja'},
    {'Firma': 'advolaw GmbH', 'Website': 'advolaw.ch', 'EMail': 'info@advolaw.ch', 'Stadt': 'Luzern', 'Telefon': '+41 41 460 55 33', 'WordPress': 'Ja'},
    {'Firma': 'Wadsack Legal GmbH', 'Website': 'wadsack-legal.ch', 'EMail': 'lukas.wadsack@wadsack.ch', 'Stadt': 'Zug', 'Telefon': '+41 41 720 30 75', 'WordPress': 'Ja'},
    {'Firma': 'Novalex Rechtsanwälte AG', 'Website': 'novalex.ch', 'EMail': 'info@novalex.ch', 'Stadt': 'Teufen', 'Telefon': '+41 71 510 92 20', 'WordPress': 'Ja'},
    {'Firma': 'thalhammer | bossart | von rohr', 'Website': 'lawsg.ch', 'EMail': 'mail@lawsg.ch', 'Stadt': 'St. Gallen', 'Telefon': '+41 71 222 23 53', 'WordPress': 'Ja'},
    {'Firma': 'AK legal (C. Kapfhamer)', 'Website': 'ak-legal.ch', 'EMail': 'kapfhamer@ak-legal.ch', 'Stadt': 'Bottighofen', 'Telefon': '+41 71 571 20 90', 'WordPress': 'Ja'},
    {'Firma': 'S-E-K Advokaten', 'Website': 's-e-k.ch', 'EMail': 'info@s-e-k.ch', 'Stadt': 'Frauenfeld', 'Telefon': '+41 52 365 11 41', 'WordPress': 'Ja'},
    {'Firma': 'Fricker Füllemann', 'Website': 'ff-law.ch', 'EMail': 'fuellemann@ff-law.ch', 'Stadt': 'Winterthur', 'Telefon': '+41 52 222 01 20', 'WordPress': 'Ja'},
    {'Firma': 'Wieduwilt Rechtsanwälte AG', 'Website': 'wieduwilt.ch', 'EMail': '', 'Stadt': 'Winterthur', 'Telefon': '+41 52 262 70 30', 'WordPress': 'Ja'},
    {'Firma': 'Anwalt Schaffhausen', 'Website': 'anwaltschaffhausen.ch', 'EMail': 'info@anwaltschaffhausen.ch', 'Stadt': 'Schaffhausen', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Anwaltskanzlei Zbinden AG', 'Website': 'anwaltskanzlei-sh.ch', 'EMail': 'info@anwaltskanzlei-sh.ch', 'Stadt': 'Schaffhausen', 'Telefon': '+41 52 630 21 00', 'WordPress': 'Ja'},
    {'Firma': 'Schiltknecht Rechtsanwälte', 'Website': 'schiltknecht-rechtsanwalt.ch', 'EMail': 'info@schiltknecht-rechtsanwalt.ch', 'Stadt': 'Aarau', 'Telefon': '+41 62 834 33 00', 'WordPress': 'Ja'},
    {'Firma': 'Geissmann Rechtsanwälte AG', 'Website': 'geissmannlegal.ch', 'EMail': 'mail@geissmannlegal.ch', 'Stadt': 'Baden', 'Telefon': '+41 56 203 00 11', 'WordPress': 'Ja'},
    {'Firma': 'BodmerFischer AG', 'Website': 'bodmerfischer.ch', 'EMail': 'info@bodmerfischer.ch', 'Stadt': 'Zürich', 'Telefon': '+41 44 711 71 71', 'WordPress': 'Ja'},
]


def normalize_url(url):
    url = str(url).strip().lower().rstrip('/')
    url = url.replace('https://', '').replace('http://', '').replace('www.', '')
    return url


wb = openpyxl.load_workbook(input_file)
ws = wb["Anwalts Kanzleien"]

existing_urls = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    url = normalize_url(row[0].value or '')
    if url:
        existing_urls.add(url)

print(f"Bestehende Eintraege: {len(existing_urls)}")

header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
data_font = Font(name='Calibri', size=11)
data_alignment = Alignment(vertical='center', wrap_text=False)

added = 0
skipped = 0

for lead in new_leads:
    url_norm = normalize_url(lead['Website'])
    if url_norm in existing_urls:
        print(f"  SKIP (Duplikat): {lead['Firma']} ({lead['Website']})")
        skipped += 1
        continue

    next_row = ws.max_row + 1
    col_map = {1: 'Firma', 2: 'Website', 3: 'EMail', 4: 'Stadt', 5: 'Telefon', 6: 'WordPress'}

    for col_idx, field in col_map.items():
        cell = ws.cell(row=next_row, column=col_idx, value=lead.get(field, ''))
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

    existing_urls.add(url_norm)
    added += 1
    print(f"  ADD: {lead['Firma']} ({lead['Website']}) - {lead['Stadt']}")

print(f"\nErgebnis: {added} neue Leads hinzugefuegt, {skipped} Duplikate uebersprungen")
print(f"Gesamtanzahl: {ws.max_row - 1} Eintraege")

wb.save(input_file)
print(f"Gespeichert: {input_file}")
