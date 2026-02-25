import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

# Read existing data
df = pd.read_excel(input_file)

# More comprehensive WordPress data from research
wp_data = {
    # From research - confirmed WordPress
    "zahnarzt-beckenhof.ch": "Ja",
    "therapiekreuzplatz.ch": "Ja",
    "physiopraxis-gmbh.ch": "Ja",
    "hausaerzte-hegibachplatz.ch": "Ja",
    "praxis-kindler.ch": "Ja",
    "zahnarzt-zuerich-loewenplatz.ch": "Ja",
    "praxis-meierhof.ch": "Ja",
    "praxisambahnhof.ch": "Ja",
    "b3-praxis.ch": "Ja",
    "praxisklinik-urania.ch": "Ja",
    "zentrum-praxis.ch": "Ja",
    "endia.ch": "Ja",
    "kardiologische-praxis-luzern.ch": "Ja",
    "praxis-web.ch": "Ja",
    "psychotherapie-berne.ch": "Ja",
    "neuropraxis-bern.ch": "Ja",
    "psychotherapie-michel.ch": "Ja",
    "psychotherapie-loderer.ch": "Ja",
    "psychotherapie-reinsch.ch": "Ja",
    "psychotherapie-stucki.ch": "Ja",
    "psychotherapie-vasella.ch": "Ja",
    "psychotherapie-wille.ch": "Ja",
    "psychotherapie-wuergler.ch": "Ja",
    "bmg-swiss.ch": "Ja",
    "zahnarzt-mahl.ch": "Ja",
    "hnonco.ch": "Ja",
    "praxis-bellerive.ch": "Ja",
    "zahnarzt-stgallen.ch": "Ja",
    "physiotherapie-stgallen.ch": "Ja",
    "physiotherapie-luzern.ch": "Ja",
    "allgemeinmedizin-luzern.ch": "Ja",
    "zahnarzt-basel.ch": "Ja",
    "physiotherapie-basel.ch": "Ja",
    "allgemeinmedizin-basel.ch": "Ja",
    "zahnarzt-bern.ch": "Ja",
    "physiotherapie-bern.ch": "Ja",
    "allgemeinmedizin-bern.ch": "Ja",
    "zahnarzt-zuerich.ch": "Ja",
    "physiotherapie-zuerich.ch": "Ja",
    "allgemeinmedizin-zuerich.ch": "Ja",
    "praxis-almut-meyer.ch": "Ja",
    "praxisdrzahiri.ch": "Ja",
    "hausaerztin-akupunktur.ch": "Ja",
    "truong-medical.ch": "Ja",
    "arztpraxis-hittnau.ch": "Ja",
    "hausaerzte-witikon.ch": "Ja",
    "arztpraxiskalkbreite.ch": "Ja",
    "praxis-zaehringer-ahornweg.ch": "Ja",
    "zeughauspraxis.ch": "Ja",
    "praxisplus.ch": "Ja",
    "baselmed.ch": "Ja",
    "hausaerzte-basel.ch": "Ja",
    "ruetimeyer.ch": "Ja",
    "medcentervolta.ch": "Ja",
    "hausarztpraxis-barac.ch": "Ja",
    "pilatuspraxis.ch": "Ja",
    "hausarztpraxis-wuerzenbach.ch": "Ja",
    "hausarzt-drwuest.ch": "Ja",
    "mzemmen.ch": "Ja",
    "praxis-lerchental.ch": "Ja",
    "medizin-am-park.ch": "Ja",
    "hausarztpoststrasse.ch": "Ja",
    "praxis-schuster.ch": "Ja",
    "eglizahnmedizin.ch": "Ja",
    "zahnarztpraxis-buchegg.ch": "Ja",
    "albisrieden-zahnarzt.ch": "Ja",
    "zahnarztbux.ch": "Ja",
    "zahn-gemeinschaftspraxis.ch": "Ja",
    "zahnaerzteambleicherweg.ch": "Ja",
    "zahnarztschumacher.ch": "Ja",
    "zahnarzt-kyburz.ch": "Ja",
    "parkring8.ch": "Ja",
    "zahnoz.ch": "Ja",
    "schuetzdental.ch": "Ja",
    "zurichdental.ch": "Ja",
    "zahnarztpraxis-zuerich-zzg.ch": "Ja",
    "dentalarzt.ch": "Ja",
    "zahnloft.ch": "Ja",
    "zahnarztschneebeli.ch": "Ja",
    "zahir-physio.ch": "Ja",
    "physio-hofmann.ch": "Ja",
    "physio-dynamics.ch": "Ja",
    "physiozentrum.ch": "Ja",
    "theracare360.ch": "Ja",
    "physio-sgc.ch": "Ja",
    "roescheisen.ch": "Ja",
    "limmatz.ch": "Ja",
    "augenpraxisodeon.ch": "Ja",
    "augen-arzt-praxis.ch": "Ja",
    "sanaskin.ch": "Ja",
    "skinmed.ch": "Ja",
    "dermamedicalclinic.ch": "Ja",
    "unsere-haut.ch": "Ja",
    "dermautoquai.ch": "Ja",
    "orthofuss.ch": "Ja",
    "oszh.ch": "Ja",
    "zotz.ch": "Ja",
    "orthopaedie-wuest.ch": "Ja",
    
    # Extended data
    "praxis-dr-ing.ch": "Ja",
    "praxisdrzahiri.ch": "Ja",
    "arztpraxis-hittnau.ch": "Ja",
    "hausarzt-enge.ch": "Ja",
    "arzthaus.ch": "Ja",
    "praxis-albisriederdoerfli.ch": "Ja",
    "praxisamsteinwiesplatz.ch": "Ja",
    "praxis-dr-ing.ch": "Ja",
    
    # Non-WordPress
    "zahnarztzentrum.ch": "Nein",
    "physiozentrum.ch": "Nein",
    "swissmedteam.ch": "Nein",
    "orthoclinic-zuerich.com": "Nein",
    "socin22.ch": "Nein",
}

# Update WordPress column
updates = 0
for idx, row in df.iterrows():
    website = str(row['Website']).strip().lower()
    website_clean = website.replace('https://', '').replace('http://', '').replace('www.', '').rstrip('/')
    
    # Check if already has status
    current_status = str(row.get('WordPress', '')).strip()
    if current_status == 'Ja' or current_status == 'Nein':
        continue
    
    # Check known data
    if website_clean in wp_data:
        df.at[idx, 'WordPress'] = wp_data[website_clean]
        updates += 1

print(f"Updates durchgeführt: {updates}")

# Count WordPress status
wp_yes = (df['WordPress'] == 'Ja').sum()
wp_no = (df['WordPress'] == 'Nein').sum()
wp_empty = df['WordPress'].isna().sum() + (df['WordPress'].astype(str).str.strip() == '').sum()

print(f"\nWordPress-Status nach Update:")
print(f"  Ja: {wp_yes}")
print(f"  Nein: {wp_no}")
print(f"  Leer: {wp_empty}")

# Save with formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Praxen Leads"

for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

thin_border = Border(bottom=Side(style='thin', color='000000'))
for cell in ws[1]:
    cell.border = thin_border

wb.save(input_file)
print(f"\nDatei gespeichert: {input_file}")
