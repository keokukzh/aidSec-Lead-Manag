import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen leads - Tabellenblatt1.csv"
output_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

# Read CSV
df = pd.read_csv(input_file, header=None)

# Split into two sections based on structure
section1 = df.iloc[1:51].copy()  # Rows 1-50 (0-indexed: 1-50)
section2 = df.iloc[52:].copy()   # Rows 52-end

# Structure 1: Firma, Website, CMS, E-Mail, Stadt, Telefon, nDSG-Check
section1.columns = ['Firma', 'Website', 'CMS', 'EMail', 'Stadt', 'Telefon', 'nDSG']
section1 = section1[['Firma', 'Website', 'EMail', 'Stadt', 'Telefon']]

# Structure 2: Praxis, Website, E-Mail, Stadt, Telefonnummer (leere Spalten danach)
section2 = section2.iloc[:, :5].copy()  # Nur erste 5 Spalten
section2.columns = ['Firma', 'Website', 'EMail', 'Stadt', 'Telefon']
section2 = section2[['Firma', 'Website', 'EMail', 'Stadt', 'Telefon']]

# Combine both sections
combined = pd.concat([section1, section2], ignore_index=True)

# Clean up - remove empty rows
combined = combined.dropna(subset=['Firma'])
combined = combined[combined['Firma'].str.strip() != '']

# Clean website URLs - remove < > brackets
combined['Website'] = combined['Website'].astype(str).str.replace('<', '').str.replace('>', '')

# Remove duplicates based on Website and EMail
before_count = len(combined)
deduplicated = combined.drop_duplicates(subset=['Website', 'EMail'], keep='first')
after_count = len(deduplicated)

print(f"Zeilen vor Duplikatentfernung: {before_count}")
print(f"Zeilen nach Duplikatentfernung: {after_count}")
print(f"Duplikate entfernt: {before_count - after_count}")

# Reset index
deduplicated = deduplicated.reset_index(drop=True)

# Create Excel with formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Praxen Leads"

# Write data
for r_idx, row in enumerate(dataframe_to_rows(deduplicated, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        # Format header row
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

# Set column widths
ws.column_dimensions['A'].width = 40  # Firma
ws.column_dimensions['B'].width = 35  # Website
ws.column_dimensions['C'].width = 35  # EMail
ws.column_dimensions['D'].width = 18  # Stadt
ws.column_dimensions['E'].width = 20  # Telefon

# Add thin borders to header
thin_border = Border(
    bottom=Side(style='thin', color='000000')
)
for cell in ws[1]:
    cell.border = thin_border

# Save
wb.save(output_file)
print(f"\nDatei gespeichert: {output_file}")
print(f"Liste enthält {after_count} eindeutige Praxen")
