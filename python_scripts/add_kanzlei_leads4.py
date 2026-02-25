import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

new_leads = [
    {'Firma': 'PALOMBO Anwaltskanzlei', 'Website': 'palombo-rechtsanwalt.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '+41 43 317 96 48', 'WordPress': 'Ja'},
    {'Firma': 'schadenanwaelte AG', 'Website': 'schadenanwaelte.ch', 'EMail': 'kanzlei@schadenanwaelte.ch', 'Stadt': 'Zürich', 'Telefon': '+41 58 252 52 52', 'WordPress': 'Ja'},
    {'Firma': 'Schmid & Herrmann Rechtsanwälte', 'Website': 'shanwaelte.ch', 'EMail': 'patrick.wagner@shanwaelte.ch', 'Stadt': 'Basel', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'KSPartner', 'Website': 'kspartner.ch', 'EMail': 'sfriedauer@kspartner.ch', 'Stadt': 'Zürich', 'Telefon': '+41 44 388 57 57', 'WordPress': 'Ja'},
    {'Firma': 'Die Advokatur.ch', 'Website': 'dieadvokatur.ch', 'EMail': '', 'Stadt': '', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'Hanhart Law', 'Website': 'hanhartlaw.ch', 'EMail': 'info@hanhart-anwalt.ch', 'Stadt': 'Zürich', 'Telefon': '+41 44 545 54 50', 'WordPress': 'Ja'},
    {'Firma': 'Advokatur Franziska Schnyder', 'Website': 'schnyder-bern.ch', 'EMail': 'advo.schnyder@bluewin.ch', 'Stadt': 'Bern', 'Telefon': '+41 31 311 26 26', 'WordPress': 'Ja'},
    {'Firma': 'Anwalt-Migrationsrecht.ch', 'Website': 'anwalt-migrationsrecht.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'SteuriFisch AG', 'Website': 'steurifisch.ch', 'EMail': '', 'Stadt': 'Wil SG', 'Telefon': '+41 71 511 45 05', 'WordPress': 'Ja'},
    {'Firma': 'Anwaltskanzlei Zürichsee AG', 'Website': 'anwaltskanzlei-zuerichsee.ch', 'EMail': 'hl@ak-see.ch', 'Stadt': 'Zürichsee', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'RA Dr. John Trachsel', 'Website': 'rechtsanwalt-trachsel.ch', 'EMail': 'mail@rechtsanwalt-trachsel.ch', 'Stadt': 'Zürich', 'Telefon': '+41 44 243 03 00', 'WordPress': 'Ja'},
    {'Firma': 'Engelberger Anwälte & Notare', 'Website': 'engelberger-anwaelte.ch', 'EMail': 'engelberger@engelberger.law', 'Stadt': 'Sarnen', 'Telefon': '+41 41 229 30 20', 'WordPress': 'Ja'},
    {'Firma': 'Anwaltskanzlei Brandon-Kaufmann', 'Website': 'anwaltskanzlei-in-zuerich.ch', 'EMail': 'beratung@anwaltskanzlei-in-zuerich.ch', 'Stadt': 'Zürich', 'Telefon': '+41 44 362 09 24', 'WordPress': 'Ja'},
    {'Firma': 'Advokatur Held', 'Website': 'advokatur-held.ch', 'EMail': '', 'Stadt': 'Basel', 'Telefon': '+41 61 554 58 85', 'WordPress': 'Ja'},
    {'Firma': 'Spühler Rechtsanwälte AG', 'Website': 'spuehler.legal', 'EMail': 'info@spuehler.legal', 'Stadt': 'Zürich', 'Telefon': '+41 43 344 05 05', 'WordPress': 'Ja'},
    {'Firma': 'RLH Law', 'Website': 'rlh-law.ch', 'EMail': 'laeuffer@rlh-law.ch', 'Stadt': 'Baden', 'Telefon': '+41 56 521 44 00', 'WordPress': 'Ja'},
]

def normalize_url(url):
    url = str(url).strip().lower().rstrip('/')
    url = url.replace('https://', '').replace('http://', '').replace('www.', '')
    return url

wb = openpyxl.load_workbook(input_file)
ws = wb['Anwalts Kanzleien']

existing_urls = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    val = str(row[0].value or '').strip()
    if val:
        existing_urls.add(normalize_url(val))

print(f"Bestehende Eintraege: {len(existing_urls)}")

thin_border = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0')
)
data_font = Font(name='Calibri', size=11, color='333333')
data_alignment = Alignment(horizontal='left', vertical='center')

added = 0
skipped = 0

for lead in new_leads:
    norm_url = normalize_url(lead['Website'])
    if norm_url in existing_urls:
        print(f"  SKIP (Duplikat): {lead['Website']}")
        skipped += 1
        continue

    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=lead['Firma'])
    ws.cell(row=next_row, column=2, value=lead['Website'])
    ws.cell(row=next_row, column=3, value=lead['EMail'])
    ws.cell(row=next_row, column=4, value=lead['Stadt'])
    ws.cell(row=next_row, column=5, value=lead['Telefon'])
    ws.cell(row=next_row, column=6, value=lead['WordPress'])

    for col in range(1, 7):
        cell = ws.cell(row=next_row, column=col)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

    existing_urls.add(norm_url)
    added += 1
    print(f"  ADDED: {lead['Firma']} ({lead['Website']})")

wb.save(input_file)
print(f"\nErgebnis: {added} hinzugefuegt, {skipped} uebersprungen")
print(f"Gesamt 'Anwalts Kanzleien': {ws.max_row - 1} Eintraege")
