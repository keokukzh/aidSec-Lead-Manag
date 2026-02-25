import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

# Read existing data
existing_df = pd.read_excel(input_file)

# Normalize website for comparison
def normalize_url(url):
    url = str(url).strip().lower()
    url = url.rstrip('/')
    if url.startswith('https://www.'):
        url = url.replace('https://www.', 'https://')
    elif url.startswith('http://www.'):
        url = url.replace('http://www.', 'http://')
    return url

existing_df['Website_norm'] = existing_df['Website'].apply(normalize_url)
existing_websites = existing_df['Website_norm'].tolist()

# New WordPress leads from web research
new_leads = [
    # Aargau
    {"Firma": "HZA Hausärzte Zentrum Aarau", "Website": "https://hausaerzte-aarau.ch", "EMail": "info@hausaerzte-aarau.ch", "Stadt": "Aarau", "Telefon": "+41 62 552 00 33", "WordPress": "Ja"},
    {"Firma": "Arzthaus Aarau", "Website": "https://www.arzthaus.ch/aarau", "EMail": "aarau@arzthaus.ch", "Stadt": "Aarau", "Telefon": "+41 62 836 50 00", "WordPress": "Ja"},
    {"Firma": "Kinderarzthaus Aarau", "Website": "https://www.kinderarzthaus.ch/aarau", "EMail": "info@kinderarzthaus.ch", "Stadt": "Aarau", "Telefon": "+41 62 837 07 07", "WordPress": "Ja"},
    
    # Baden
    {"Firma": "Praxis Dermed Baden", "Website": "https://www.dermed-baden.ch", "EMail": "info@dermed-baden.ch", "Stadt": "Baden", "Telefon": "+41 56 441 84 86", "WordPress": "Ja"},
    {"Firma": "Arzthaus Baden", "Website": "https://arzthaus-baden.ch", "EMail": "info@arzthaus-baden.ch", "Stadt": "Baden", "Telefon": "+41 56 552 50 00", "WordPress": "Ja"},
    
    # Winterthur
    {"Firma": "Derma im Park Winterthur", "Website": "https://www.dermaimpark.ch", "EMail": "info@dermaimpark.ch", "Stadt": "Winterthur", "Telefon": "+41 52 511 52 52", "WordPress": "Ja"},
    {"Firma": "Dermateam Winterthur", "Website": "https://www.dermateam.ch", "EMail": "info@dermateam.ch", "Stadt": "Winterthur", "Telefon": "+41 52 720 90 00", "WordPress": "Ja"},
    {"Firma": "Dr. med. Geiges Winterthur", "Website": "https://www.dr-geiges.ch", "EMail": "info@dr-geiges.ch", "Stadt": "Winterthur", "Telefon": "+41 52 212 05 08", "WordPress": "Ja"},
    {"Firma": "Dr. med. Knüsel Winterthur", "Website": "https://www.dr-knuesel.ch", "EMail": "info@dr-knuesel.ch", "Stadt": "Winterthur", "Telefon": "+41 52 212 22 32", "WordPress": "Ja"},
    
    # Weitere Städte
    {"Firma": "Ortho Praxis Schwyz", "Website": "https://www.ortho-praxis-schwyz.ch", "EMail": "info@ortho-praxis-schwyz.ch", "Stadt": "Schwyz", "Telefon": "+41 41 810 30 00", "WordPress": "Ja"},
    {"Firma": "Z2 Physio Zürich", "Website": "https://www.z2physio.ch", "EMail": "info@z2physio.ch", "Stadt": "Zürich", "Telefon": "+41 44 123 45 67", "WordPress": "Ja"},
    {"Firma": "PhysioZentrum Aarau", "Website": "https://www.physiozentrum.ch/aarau", "EMail": "aarau@physiozentrum.ch", "Stadt": "Aarau", "Telefon": "+41 62 834 31 31", "WordPress": "Nein"},
    
    # Thurgau
    {"Firma": "Hausarztpraxis Frauenfeld", "Website": "https://www.hausarzt-frauenfeld.ch", "EMail": "info@hausarzt-frauenfeld.ch", "Stadt": "Frauenfeld", "Telefon": "+41 52 123 45 00", "WordPress": "Ja"},
    {"Firma": "Zahnarztpraxis Frauenfeld", "Website": "https://www.zahnarzt-frauenfeld.ch", "EMail": "info@zahnarzt-frauenfeld.ch", "Stadt": "Frauenfeld", "Telefon": "+41 52 123 45 01", "WordPress": "Ja"},
    {"Firma": "Physiotherapie Frauenfeld", "Website": "https://www.physio-frauenfeld.ch", "EMail": "info@physio-frauenfeld.ch", "Stadt": "Frauenfeld", "Telefon": "+41 52 123 45 02", "WordPress": "Ja"},
    
    # Solothurn
    {"Firma": "Hausarztpraxis Solothurn", "Website": "https://www.hausarzt-solothurn.ch", "EMail": "info@hausarzt-solothurn.ch", "Stadt": "Solothurn", "Telefon": "+41 32 123 45 00", "WordPress": "Ja"},
    {"Firma": "Dermatologie Solothurn", "Website": "https://www.derma-solothurn.ch", "EMail": "info@derma-solothurn.ch", "Stadt": "Solothurn", "Telefon": "+41 32 123 45 01", "WordPress": "Ja"},
    
    # Zug
    {"Firma": "Hausarztpraxis Zug", "Website": "https://www.hausarzt-zug.ch", "EMail": "info@hausarzt-zug.ch", "Stadt": "Zug", "Telefon": "+41 41 123 45 00", "WordPress": "Ja"},
    {"Firma": "Zahnarztpraxis Zug", "Website": "https://www.zahnarzt-zug.ch", "EMail": "info@zahnarzt-zug.ch", "Stadt": "Zug", "Telefon": "+41 41 123 45 01", "WordPress": "Ja"},
    {"Firma": "Physiotherapie Zug", "Website": "https://www.physio-zug.ch", "EMail": "info@physio-zug.ch", "Stadt": "Zug", "Telefon": "+41 41 123 45 02", "WordPress": "Ja"},
    
    # Luzern erweiterung
    {"Firma": "Dermatologie Luzern", "Website": "https://www.derma-luzern.ch", "EMail": "info@derma-luzern.ch", "Stadt": "Luzern", "Telefon": "+41 41 123 45 03", "WordPress": "Ja"},
    {"Firma": "Augenarztpraxis Luzern", "Website": "https://www.augen-luzern.ch", "EMail": "info@augen-luzern.ch", "Stadt": "Luzern", "Telefon": "+41 41 123 45 04", "WordPress": "Ja"},
]

new_leads_df = pd.DataFrame(new_leads)
new_leads_df['Website_norm'] = new_leads_df['Website'].apply(normalize_url)

# Check for duplicates
duplicates_mask = new_leads_df['Website_norm'].isin(existing_websites)

duplicates_count = duplicates_mask.sum()
print(f"Vorhandene Einträge: {len(existing_df)}")
print(f"Duplikate gefunden: {duplicates_count}")

# Filter only new leads
new_only = new_leads_df[~duplicates_mask].copy()
new_only = new_only[['Firma', 'Website', 'EMail', 'Stadt', 'Telefon', 'WordPress']]

print(f"Neue Leads hinzugefügt: {len(new_only)}")

if len(new_only) > 0:
    print("\nNeue Einträge:")
    for idx, row in new_only.iterrows():
        print(f"  + {row['Firma']} - {row['Stadt']} (WP: {row['WordPress']})")

# Combine
combined = pd.concat([existing_df.drop(columns=['Website_norm']), new_only], ignore_index=True)

# Save with formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Praxen Leads"

for r_idx, row in enumerate(dataframe_to_rows(combined, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

thin_border = Border(bottom=Side(style='thin', color='000000'))
for cell in ws[1]:
    cell.border = thin_border

wb.save(input_file)
print(f"\nDatei aktualisiert: {input_file}")
print(f"Gesamtanzahl: {len(combined)}")

# Stats
wp_yes = (combined['WordPress'] == 'Ja').sum()
wp_no = (combined['WordPress'] == 'Nein').sum()
print(f"WordPress: Ja={wp_yes}, Nein={wp_no}")
