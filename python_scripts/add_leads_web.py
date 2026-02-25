import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from urllib.parse import urlparse

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

# Read existing data
existing_df = pd.read_excel(input_file)

# Normalize website for comparison
def normalize_url(url):
    url = str(url).strip().lower()
    url = url.rstrip('/')
    if url.startswith('https://www.'):
        url = url.replace('https://www.', 'https://')
    elif url.startswith('http://www.'):
        url = url.replace('http://www.', 'http://')
    return url

existing_df['Website_norm'] = existing_df['Website'].apply(normalize_url)
existing_websites = existing_df['Website_norm'].tolist()

# New leads from web research
new_leads = [
    {"Firma": "Praxis Seefeld", "Website": "https://praxis-seefeld.ch", "EMail": "", "Stadt": "Zürich", "Telefon": ""},
    {"Firma": "Praxis Altstetten", "Website": "https://www.praxis-altstetten.ch", "EMail": "", "Stadt": "Zürich", "Telefon": ""},
    {"Firma": "Praxis Goldbrunnen", "Website": "https://www.praxis-goldbrunnen.ch", "EMail": "", "Stadt": "Zürich", "Telefon": ""},
    {"Firma": "Zahnarztzentrum.ch", "Website": "https://zahnarztzentrum.ch", "EMail": "info@zahnarztzentrum.ch", "Stadt": "Bern", "Telefon": ""},
    {"Firma": "Zahnarztzentrum.ch Basel", "Website": "https://zahnarztzentrum.ch", "EMail": "info@zahnarztzentrum.ch", "Stadt": "Basel", "Telefon": ""},
    {"Firma": "Zahnarztzentrum.ch Luzern", "Website": "https://zahnarztzentrum.ch", "EMail": "info@zahnarztzentrum.ch", "Stadt": "Luzern", "Telefon": ""},
    {"Firma": "Zahnarzt in Luzern", "Website": "https://zahnarzt-in-luzern.ch", "EMail": "", "Stadt": "Luzern", "Telefon": ""},
    {"Firma": "Zahnarzt in Basel", "Website": "https://www.zahnarzt-in-basel.ch", "EMail": "", "Stadt": "Basel", "Telefon": ""},
    {"Firma": "Praxis Salzmann", "Website": "https://www.praxis-salzmann.ch", "EMail": "info@praxis-salzmann.ch", "Stadt": "Zürich", "Telefon": "+41 44 350 63 00"},
    {"Firma": "Kinezi Sport Physiotherapie", "Website": "https://www.sport-physiotherapie-zuerich.ch", "EMail": "", "Stadt": "Zürich", "Telefon": ""},
    {"Firma": "HautZentrum Zürich", "Website": "https://www.hautzentrum-zuerich.ch", "EMail": "", "Stadt": "Zürich", "Telefon": ""},
    {"Firma": "Augenarzt Zürich Schaffhauserplatz", "Website": "https://www.augenaerzte-zuerich.ch", "EMail": "", "Stadt": "Zürich", "Telefon": "+41 43 333 44 44"},
    {"Firma": "Augenarztpraxis Hohermuth & Rüegg", "Website": "https://augen-arzt-praxis.ch", "EMail": "", "Stadt": "Zürich", "Telefon": ""},
    {"Firma": "zmed Ärztenetz", "Website": "https://www.zmed.ch", "EMail": "", "Stadt": "Zürich", "Telefon": ""},
]

new_leads_df = pd.DataFrame(new_leads)
new_leads_df['Website_norm'] = new_leads_df['Website'].apply(normalize_url)

# Check for duplicates
duplicates_mask = new_leads_df['Website_norm'].isin(existing_websites)

duplicates_count = duplicates_mask.sum()
print(f"Vorhandene Einträge: {len(existing_df)}")
print(f"Duplikate gefunden: {duplicates_count}")

# Show duplicates
if duplicates_count > 0:
    print("\nDuplikate (werden übersprungen):")
    for idx, row in new_leads_df[duplicates_mask].iterrows():
        print(f"  - {row['Firma']} ({row['Website']})")

# Filter only new leads
new_only = new_leads_df[~duplicates_mask].copy()
new_only = new_only[['Firma', 'Website', 'EMail', 'Stadt', 'Telefon']]

print(f"Neue Leads hinzugefügt: {len(new_only)}")

# Show new leads
if len(new_only) > 0:
    print("\nNeue Einträge:")
    for idx, row in new_only.iterrows():
        print(f"  + {row['Firma']} - {row['Stadt']}")

# Combine
combined = pd.concat([existing_df.drop(columns=['Website_norm']), new_only], ignore_index=True)

# Save with formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Praxen Leads"

for r_idx, row in enumerate(dataframe_to_rows(combined, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

ws.column_dimensions['A'].width = 40
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 35
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 20

thin_border = Border(bottom=Side(style='thin', color='000000'))
for cell in ws[1]:
    cell.border = thin_border

wb.save(input_file)
print(f"\nDatei aktualisiert: {input_file}")
print(f"Gesamtanzahl Einträge: {len(combined)}")
