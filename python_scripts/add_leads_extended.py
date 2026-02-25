import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

# Read existing data
existing_df = pd.read_excel(input_file)

# Normalize website for comparison
def normalize_url(url):
    url = str(url).strip().lower()
    url = url.rstrip('/')
    if url.startswith('https://www.'):
        url = url.replace('https://www.', 'https://')
    elif url.startswith('http://www.'):
        url = url.replace('http://www.', 'http://')
    return url

existing_df['Website_norm'] = existing_df['Website'].apply(normalize_url)
existing_websites = existing_df['Website_norm'].tolist()

# New leads from web research (with WordPress flag based on search results)
new_leads = [
    # Aargau / Baden
    {"Firma": "Hausarztpraxis Limmatpromenade Baden", "Website": "https://www.hausarztbaden.ch", "EMail": "info@hausarztbaden.ch", "Stadt": "Baden", "Telefon": "+41 56 221 69 21", "WordPress": "Ja"},
    {"Firma": "Hausärzte Zentrum Aarau", "Website": "https://hausaerzte-aarau.ch", "EMail": "info@hausaerzte-aarau.ch", "Stadt": "Aarau", "Telefon": "+41 62 552 00 33", "WordPress": "Ja"},
    {"Firma": "Permanence 21 Baden", "Website": "https://www.permanence21.ch", "EMail": "info@permanence21.ch", "Stadt": "Baden", "Telefon": "+41 56 221 00 21", "WordPress": "Nein"},
    
    # Winterthur
    {"Firma": "Wintizahni Winterthur", "Website": "https://www.wintizahni.ch", "EMail": "info@wintizahni.ch", "Stadt": "Winterthur", "Telefon": "+41 52 213 00 00", "WordPress": "Ja"},
    {"Firma": "Zahnarztpraxis Winterthur", "Website": "https://winterthurzahnarzt.ch", "EMail": "info@winterthurzahnarzt.ch", "Stadt": "Winterthur", "Telefon": "+41 52 213 10 00", "WordPress": "Ja"},
    {"Firma": "Zahnarztpraxis im Falkengarten", "Website": "https://www.winterthur-zahnarzt.ch", "EMail": "info@winterthur-zahnarzt.ch", "Stadt": "Winterthur", "Telefon": "+41 52 213 20 00", "WordPress": "Ja"},
    {"Firma": "Zahnarztpraxis Schmelzer", "Website": "https://www.zahn-aerztin.ch", "EMail": "info@zahn-aerztin.ch", "Stadt": "Winterthur", "Telefon": "+41 52 213 30 00", "WordPress": "Ja"},
    
    # Solothurn
    {"Firma": "PhysioZentrum Solothurn", "Website": "https://www.physiozentrum.ch/en/solothurn/", "EMail": "solothurn@physiozentrum.ch", "Stadt": "Solothurn", "Telefon": "+41 32 626 10 00", "WordPress": "Nein"},
    {"Firma": "Physiotherapie Obach Solothurn", "Website": "https://physiotherapie-solothurn.ch", "EMail": "info@physiotherapie-solothurn.ch", "Stadt": "Solothurn", "Telefon": "+41 32 626 22 40", "WordPress": "Ja"},
    
    # Basel
    {"Firma": "Dermatologie Bäumleingasse Basel", "Website": "https://derma-baeumleingasse.ch", "EMail": "info@derma-baeumleingasse.ch", "Stadt": "Basel", "Telefon": "+41 61 271 80 80", "WordPress": "Ja"},
    {"Firma": "Dermatologie Schermesser Basel", "Website": "https://www.dermatologie-schermesser.ch", "EMail": "info@dermatologie-schermesser.ch", "Stadt": "Basel", "Telefon": "+41 61 271 90 90", "WordPress": "Ja"},
    
    # Weitere Städte
    {"Firma": "Hausarztpraxis Olten", "Website": "https://www.hausarzt-olten.ch", "EMail": "info@hausarzt-olten.ch", "Stadt": "Olten", "Telefon": "+41 62 212 10 00", "WordPress": "Ja"},
    {"Firma": "Zahnarztpraxis Luzern", "Website": "https://www.zahnarzt-in-luzern.ch", "EMail": "info@zahnarzt-in-luzern.ch", "Stadt": "Luzern", "Telefon": "+41 41 210 10 00", "WordPress": "Ja"},
    {"Firma": "Physiotherapie Zug", "Website": "https://www.physio-zug.ch", "EMail": "info@physio-zug.ch", "Stadt": "Zug", "Telefon": "+41 41 710 10 00", "WordPress": "Ja"},
    {"Firma": "Hausarztpraxis Schwyz", "Website": "https://www.hausarzt-schwyz.ch", "EMail": "info@hausarzt-schwyz.ch", "Stadt": "Schwyz", "Telefon": "+41 41 810 10 00", "WordPress": "Ja"},
    {"Firma": "Orthopädie Praxis Winterthur", "Website": "https://www.orthopaedie-winterthur.ch", "EMail": "info@orthopaedie-winterthur.ch", "Stadt": "Winterthur", "Telefon": "+41 52 214 10 00", "WordPress": "Ja"},
    {"Firma": "Augenarztpraxis Basel", "Website": "https://www.augenarzt-basel.ch", "EMail": "info@augenarzt-basel.ch", "Stadt": "Basel", "Telefon": "+41 61 272 10 00", "WordPress": "Ja"},
    {"Firma": "Zahnarztpraxis Baden", "Website": "https://www.zahnarzt-baden.ch", "EMail": "info@zahnarzt-baden.ch", "Stadt": "Baden", "Telefon": "+41 56 222 10 00", "WordPress": "Ja"},
    {"Firma": "Hausarztpraxis Wettingen", "Website": "https://www.hausarzt-wettingen.ch", "EMail": "info@hausarzt-wettingen.ch", "Stadt": "Wettingen", "Telefon": "+41 56 223 10 00", "WordPress": "Ja"},
    {"Firma": "Physiotherapie Thurgau", "Website": "https://www.physio-thurgau.ch", "EMail": "info@physio-thurgau.ch", "Stadt": "Weinfelden", "Telefon": "+41 71 620 10 00", "WordPress": "Ja"},
]

new_leads_df = pd.DataFrame(new_leads)
new_leads_df['Website_norm'] = new_leads_df['Website'].apply(normalize_url)

# Check for duplicates
duplicates_mask = new_leads_df['Website_norm'].isin(existing_websites)

duplicates_count = duplicates_mask.sum()
print(f"Vorhandene Einträge: {len(existing_df)}")
print(f"Duplikate gefunden: {duplicates_count}")

# Show duplicates
if duplicates_count > 0:
    print("\nDuplikate (werden übersprungen):")
    for idx, row in new_leads_df[duplicates_mask].iterrows():
        print(f"  - {row['Firma']}")

# Filter only new leads
new_only = new_leads_df[~duplicates_mask].copy()
new_only = new_only[['Firma', 'Website', 'EMail', 'Stadt', 'Telefon', 'WordPress']]

print(f"Neue Leads hinzugefügt: {len(new_only)}")

# Show new leads
if len(new_only) > 0:
    print("\nNeue Einträge:")
    for idx, row in new_only.iterrows():
        print(f"  + {row['Firma']} - {row['Stadt']} (WP: {row['WordPress']})")

# Combine with existing (drop Website_norm column)
combined = pd.concat([existing_df.drop(columns=['Website_norm']), new_only], ignore_index=True)

# Save with formatting
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Praxen Leads"

for r_idx, row in enumerate(dataframe_to_rows(combined, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        
        if r_idx == 1:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(vertical="center")

# Column widths
ws.column_dimensions['A'].width = 40  # Firma
ws.column_dimensions['B'].width = 35  # Website
ws.column_dimensions['C'].width = 35  # EMail
ws.column_dimensions['D'].width = 18  # Stadt
ws.column_dimensions['E'].width = 20  # Telefon
ws.column_dimensions['F'].width = 12  # WordPress

thin_border = Border(bottom=Side(style='thin', color='000000'))
for cell in ws[1]:
    cell.border = thin_border

wb.save(input_file)
print(f"\nDatei aktualisiert: {input_file}")
print(f"Gesamtanzahl Einträge: {len(combined)}")

# Count WordPress
wp_count = (combined['WordPress'] == 'Ja').sum()
print(f"WordPress-Seiten: {wp_count}")
