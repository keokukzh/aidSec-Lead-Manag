import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

input_file = r"c:\Users\aidevelo\Desktop\Praxen leads\praxen_leads_bereinigt.xlsx"

new_leads = [
    # Schulrecht/Bildungsrecht
    {'Firma': 'Dr. Heinze & Partner', 'Website': 'heinze-rechtsanwaelte.ch', 'EMail': 'info@heinze-pruefungsanfechtung.ch', 'Stadt': 'Wollerau', 'Telefon': '+41 58 25 52 510', 'WordPress': 'Ja'},
    {'Firma': 'Fachstelle Schulrecht GmbH', 'Website': 'schulrecht.ch', 'EMail': '', 'Stadt': '', 'Telefon': '', 'WordPress': 'Ja'},

    # Bank/Finanzmarkt
    {'Firma': 'Bürgi Nägeli Rechtsanwälte', 'Website': 'bnlawyers.ch', 'EMail': '', 'Stadt': 'Zürich', 'Telefon': '', 'WordPress': 'Ja'},

    # Schwyz Pfäffikon
    {'Firma': 'Advokatur Bürgi & Partners', 'Website': 'buergi-chambers.com', 'EMail': '', 'Stadt': 'Pfäffikon SZ', 'Telefon': '', 'WordPress': 'Ja'},

    # Oberaargau/Langenthal
    {'Firma': 'Notar Langenthal', 'Website': 'notar-langenthal.ch', 'EMail': '', 'Stadt': 'Langenthal', 'Telefon': '', 'WordPress': 'Ja'},
    {'Firma': 'AD!VOCATE Advokatur', 'Website': 'ad-vocate.ch', 'EMail': '', 'Stadt': 'Langenthal', 'Telefon': '', 'WordPress': 'Ja'},
]

def normalize_url(url):
    url = str(url).strip().lower().rstrip('/')
    url = url.replace('https://', '').replace('http://', '').replace('www.', '')
    return url

wb = openpyxl.load_workbook(input_file)
ws = wb['Anwalts Kanzleien']

existing_urls = set()
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    val = str(row[0].value or '').strip()
    if val:
        existing_urls.add(normalize_url(val))

print(f"Bestehende Eintraege: {len(existing_urls)}")

thin_border = Border(
    left=Side(style='thin', color='E0E0E0'),
    right=Side(style='thin', color='E0E0E0'),
    top=Side(style='thin', color='E0E0E0'),
    bottom=Side(style='thin', color='E0E0E0')
)
data_font = Font(name='Calibri', size=11, color='333333')
data_alignment = Alignment(horizontal='left', vertical='center')

added = 0
skipped = 0

for lead in new_leads:
    norm_url = normalize_url(lead['Website'])
    if norm_url in existing_urls:
        print(f"  SKIP: {lead['Website']}")
        skipped += 1
        continue

    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1, value=lead['Firma'])
    ws.cell(row=next_row, column=2, value=lead['Website'])
    ws.cell(row=next_row, column=3, value=lead['EMail'])
    ws.cell(row=next_row, column=4, value=lead['Stadt'])
    ws.cell(row=next_row, column=5, value=lead['Telefon'])
    ws.cell(row=next_row, column=6, value=lead['WordPress'])

    for col in range(1, 7):
        cell = ws.cell(row=next_row, column=col)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border

    existing_urls.add(norm_url)
    added += 1
    print(f"  ADDED: {lead['Firma']} ({lead['Website']})")

wb.save(input_file)
print(f"\nErgebnis: {added} hinzugefuegt, {skipped} uebersprungen")
print(f"Gesamt 'Anwalts Kanzleien': {ws.max_row - 1} Eintraege")
